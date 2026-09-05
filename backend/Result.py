import datetime
import io
import re
import os
import sqlite3
import secrets
import smtplib
import ssl
import logging
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from contextlib import asynccontextmanager
from typing import List, Optional, Dict, Any, Tuple
from collections import Counter
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Query, Form
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from pydantic import BaseModel, ConfigDict, Field, field_validator
from sqlalchemy import create_engine, Column, Integer, String, Float, Boolean, DateTime, ForeignKey, UniqueConstraint, func, event, text, inspect
from sqlalchemy.orm import sessionmaker, Session, declarative_base, relationship
from sqlalchemy.exc import IntegrityError
import bcrypt
import jwt
import numpy as np
import pandas as pd
import pdfplumber
import openpyxl
import json
import urllib.request
import urllib.error
from dotenv import load_dotenv

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Always load .env from backend directory first, then fallback to current working directory
load_dotenv(os.path.join(BASE_DIR, ".env"))
load_dotenv()

# ==========================================
# CONFIGURATION & PATHS
# ==========================================
SECRET_KEY = os.getenv("SECRET_KEY", "GRADE_SECRET_KEY_REPLACE")
REFRESH_SECRET_KEY = os.getenv("REFRESH_SECRET_KEY", "REFRESH_GRADE_KEY")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30
REFRESH_TOKEN_EXPIRE_DAYS = 7

BATCHES_DIR = os.path.join(BASE_DIR, "batches")
os.makedirs(BATCHES_DIR, exist_ok=True)

SYSTEM_DB_URL = f"sqlite:///{os.path.join(BASE_DIR, 'system.db')}"
SUBJECTS_DB_URL = f"sqlite:///{os.path.join(BASE_DIR, 'subjects.db')}"

# ──────────────────────────────────────────
# SQLite PERFORMANCE FACTORY
# Applies WAL mode, 32 MB page cache, NORMAL
# sync, memory temp store, and 256 MB mmap
# to every connection on every engine.
# ──────────────────────────────────────────
def _make_sqlite_engine(url: str, **kwargs):
    """Create a SQLite engine with performance PRAGMAs pre-applied."""
    eng = create_engine(url, connect_args={"check_same_thread": False}, **kwargs)
    @event.listens_for(eng, "connect")
    def _set_pragmas(conn, _):
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA cache_size=-32768")    # 32 MB page cache
        conn.execute("PRAGMA synchronous=NORMAL")   # safe + fast (not FULL)
        conn.execute("PRAGMA temp_store=MEMORY")
        conn.execute("PRAGMA mmap_size=268435456")  # 256 MB memory-mapped IO
    return eng

system_engine = _make_sqlite_engine(SYSTEM_DB_URL)
SystemSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=system_engine)
SystemBase = declarative_base()

subjects_engine = _make_sqlite_engine(SUBJECTS_DB_URL)
SubjectsSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=subjects_engine)
SubjectsBase = declarative_base()

BatchBase = declarative_base()

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="auth/login")
logger = logging.getLogger("grade_system")

# ------------------------------------------
# Report-card OTP configuration
# ------------------------------------------
REPORT_CARD_TOKEN_SECRET = os.getenv("REPORT_CARD_TOKEN_SECRET", SECRET_KEY + "_report_card_otp")
REPORT_CARD_TOKEN_EXPIRE_MINUTES = 15

OTP_LENGTH = 6
OTP_EXPIRE_MINUTES = 2
OTP_MAX_ATTEMPTS = 3
OTP_RESEND_COOLDOWN_SECONDS = 10
OTP_MAX_REQUESTS_PER_HOUR = 100

SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_USE_SSL = os.getenv("SMTP_USE_SSL", "false").lower() == "true"
SMTP_FROM_EMAIL = os.getenv("SMTP_FROM_EMAIL", SMTP_USERNAME or "no-reply@ptuniv.edu.in")
SMTP_FROM_NAME = os.getenv("SMTP_FROM_NAME", "PTU Grade Portal")
SMTP_CONFIGURED = bool(SMTP_HOST and SMTP_USERNAME and SMTP_PASSWORD)

# HTTPS Email APIs (port 443 — works seamlessly on Render free tier where SMTP ports 25/465/587 are blocked)
RESEND_API_KEY = os.getenv("RESEND_API_KEY", "")
BREVO_API_KEY = os.getenv("BREVO_API_KEY", "")
SENDGRID_API_KEY = os.getenv("SENDGRID_API_KEY", "")

GRADE_POINTS = {
    "O": 10, "A+": 9, "A": 8, "B+": 7, "B": 6, "C": 5, "D": 4,
    "F": 0, "AB": 0, "ABSENT": 0, "S": 0, "E": 0, "Z": 0, "P": 0,
    # Legacy lowercase variant (kept for any existing DB rows)
    "Ab": 0,
}

# Grades that mean the student did not pass (stored as "F" after normalisation)
_FAIL_GRADES = {"F", "AB", "ABSENT", "NC", "E", "Z", "S", "P"}

def normalize_grade(raw: str) -> str:
    """
    Normalize a raw grade string for storage.
    - Absent markers (AB, ABSENT, Ab …) are stored as 'F' so all downstream
      code only needs to check for the single canonical failing grade.
    - Existing F grades are returned unchanged.
    - Valid passing grades (O, A+, A, B+, B, C, D) are returned as-is (uppercased).
    """
    if not raw:
        return "F"
    clean = raw.strip().upper()
    if clean in ("AB", "ABSENT"):
        return "F"
    # Re-map any other zero-point special grades to F for consistent storage
    if clean in _FAIL_GRADES:
        return "F"
    # Return the canonical grade letter (preserve + suffix)
    normalized = re.sub(r'[^A-Z+]', '', clean)
    return normalized if normalized else "F"

# ==========================================
# DATABASE MODELS
# ==========================================

# 1. System DB Models (Users, Resources & OTPs)
class User(SystemBase):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True)
    email = Column(String, unique=True, index=True, nullable=True)
    hashed_password = Column(String)
    role = Column(String, default="viewer")

class Resource(SystemBase):
    __tablename__ = "resources"
    id = Column(Integer, primary_key=True, index=True, autoincrement=True)
    name = Column(String, nullable=False, index=True)
    email = Column(String, unique=True, index=True, nullable=False)
    account_type = Column(String, nullable=False)
    created_at = Column(DateTime, nullable=False, default=datetime.datetime.utcnow)

class ReportCardOTP(SystemBase):
    __tablename__ = "report_card_otps"
    id = Column(Integer, primary_key=True, index=True)
    reg_no = Column(String, nullable=False, index=True)
    email = Column(String, nullable=False)
    otp_hash = Column(String, nullable=False)
    created_at = Column(DateTime, nullable=False, default=datetime.datetime.utcnow)
    expires_at = Column(DateTime, nullable=False)
    attempts = Column(Integer, nullable=False, default=0)
    consumed = Column(Boolean, nullable=False, default=False)

class StaffOTP(SystemBase):
    """OTP records for staff registration and password reset."""
    __tablename__ = "staff_otps"
    id = Column(Integer, primary_key=True, index=True)
    email = Column(String, nullable=False, index=True)
    purpose = Column(String, nullable=False)        # 'register' or 'reset'
    otp_hash = Column(String, nullable=False)
    verified = Column(Boolean, nullable=False, default=False)
    attempts = Column(Integer, nullable=False, default=0)
    created_at = Column(DateTime, nullable=False, default=datetime.datetime.utcnow)
    expires_at = Column(DateTime, nullable=False)


# 2. Subjects DB Models (University Subjects Master)
class Subject(SubjectsBase):
    __tablename__ = "subjects"
    id = Column(Integer, primary_key=True, index=True)
    code = Column(String, unique=True, index=True, nullable=False)
    name = Column(String, nullable=False)
    credits = Column(Float, nullable=False)
    semester = Column(String)
    department = Column(String, nullable=True)

# 3. Batch DB Models (Per-Batch Isolated Students & Results)
class Student(BatchBase):
    __tablename__ = "students"
    id = Column(Integer, primary_key=True, index=True)
    reg_no = Column(String, unique=True, index=True, nullable=False)
    name = Column(String, nullable=False)
    department = Column(String, nullable=False)
    programme = Column(String, nullable=True)
    batch = Column(String)
    section = Column(String, nullable=True)
    is_repeater = Column(Boolean, default=False, nullable=False)
    source = Column(String, nullable=False, default="roster")
    email = Column(String, nullable=True)
    results = relationship("Result", back_populates="student", cascade="all, delete-orphan")

class Result(BatchBase):
    __tablename__ = "results"
    id = Column(Integer, primary_key=True, index=True)
    student_id = Column(Integer, ForeignKey("students.id"), nullable=False)
    subject_id = Column(Integer, nullable=False, index=True)  # References Subject.id from subjects.db
    semester = Column(String, nullable=False)
    year = Column(String, nullable=True)
    grade = Column(String, nullable=False)
    grade_point = Column(Float, nullable=False)
    batch = Column(String, nullable=True)
    attempt = Column(Integer, default=1, nullable=False)
    had_arrear = Column(Boolean, default=False, nullable=False)
    student = relationship("Student", back_populates="results")

    __table_args__ = (UniqueConstraint('student_id', 'subject_id', 'attempt', name='uix_student_subject_attempt'),)

# ==========================================
# BATCH DATABASE MANAGER
# ==========================================
_batch_engines: Dict[str, Any] = {}
_batch_sessionmakers: Dict[str, Any] = {}

def sanitize_batch_name(batch: Optional[str]) -> str:
    raw = (batch or "").strip()
    if not raw:
        return "General"
    cleaned = re.sub(r'[^a-zA-Z0-9_\-]', '_', raw)
    return cleaned or "General"

def get_batch_db_path(batch: Optional[str]) -> str:
    sanitized = sanitize_batch_name(batch)
    return os.path.join(BATCHES_DIR, f"batch_{sanitized}.db")

def get_batch_engine(batch: Optional[str]):
    sanitized = sanitize_batch_name(batch)
    if sanitized not in _batch_engines:
        db_path = os.path.join(BATCHES_DIR, f"batch_{sanitized}.db")
        eng = _make_sqlite_engine(f"sqlite:///{db_path}")
        BatchBase.metadata.create_all(bind=eng)
        try:
            with eng.connect() as conn:
                row = conn.execute(text("SELECT sql FROM sqlite_master WHERE type='table' AND name='results'")).fetchone()
                if row and row[0] and "uix_student_subject UNIQUE" in row[0]:
                    conn.execute(text("PRAGMA foreign_keys=off"))
                    conn.execute(text("""
                        CREATE TABLE results_new (
                            id INTEGER NOT NULL PRIMARY KEY,
                            student_id INTEGER NOT NULL,
                            subject_id INTEGER NOT NULL,
                            semester VARCHAR NOT NULL,
                            year VARCHAR,
                            grade VARCHAR NOT NULL,
                            grade_point FLOAT NOT NULL,
                            batch VARCHAR,
                            attempt INTEGER DEFAULT 1 NOT NULL,
                            had_arrear BOOLEAN NOT NULL,
                            CONSTRAINT uix_student_subject_attempt UNIQUE (student_id, subject_id, attempt),
                            FOREIGN KEY(student_id) REFERENCES students (id)
                        )
                    """))
                    conn.execute(text("""
                        INSERT INTO results_new (id, student_id, subject_id, semester, year, grade, grade_point, batch, attempt, had_arrear)
                        SELECT id, student_id, subject_id, semester, year, grade, grade_point, batch, 1, had_arrear FROM results
                    """))
                    conn.execute(text("DROP TABLE results"))
                    conn.execute(text("ALTER TABLE results_new RENAME TO results"))
                    conn.execute(text("CREATE INDEX IF NOT EXISTS ix_results_id ON results (id)"))
                    conn.execute(text("CREATE INDEX IF NOT EXISTS ix_results_subject_id ON results (subject_id)"))
                    conn.execute(text("PRAGMA foreign_keys=on"))
                    conn.commit()
                else:
                    rows = conn.execute(text("PRAGMA table_info(results)")).fetchall()
                    cols = [r[1] for r in rows]
                    if cols and 'attempt' not in cols:
                        conn.execute(text("ALTER TABLE results ADD COLUMN attempt INTEGER DEFAULT 1 NOT NULL"))
                        conn.commit()
        except Exception as e:
            logger.warning("Batch DB migration notice: %s", e)
        _batch_engines[sanitized] = eng
        _batch_sessionmakers[sanitized] = sessionmaker(autocommit=False, autoflush=False, bind=eng)
    return _batch_engines[sanitized]

def get_batch_session(batch: Optional[str]) -> Session:
    sanitized = sanitize_batch_name(batch)
    get_batch_engine(sanitized)
    return _batch_sessionmakers[sanitized]()

def get_all_batch_names() -> List[str]:
    """List all batches from batch DB files — uses sqlite3 direct reads (no new ORM engine)."""
    batches = set()
    if not os.path.exists(BATCHES_DIR):
        return []
    for fname in os.listdir(BATCHES_DIR):
        if not (fname.startswith("batch_") and fname.endswith(".db")):
            continue
        raw_key = fname[len("batch_"):-len(".db")]
        if raw_key in _batch_engines:
            # Reuse existing cached session for zero overhead
            try:
                with _batch_sessionmakers[raw_key]() as s:
                    for (b,) in s.query(Student.batch).distinct().all():
                        if b and b.strip():
                            batches.add(b.strip())
            except Exception:
                pass
        else:
            # Fast direct sqlite3 read — no SQLAlchemy ORM overhead
            db_path = os.path.join(BATCHES_DIR, fname)
            try:
                conn = sqlite3.connect(db_path)
                for (b,) in conn.execute(
                    "SELECT DISTINCT batch FROM students WHERE batch IS NOT NULL AND batch != ''"
                ).fetchall():
                    if b and b.strip():
                        batches.add(b.strip())
                conn.close()
            except Exception:
                pass
    return sorted(batches)

# ──────────────────────────────────────────
# REG_NO → BATCH INDEX  (O(1) student lookup)
# ──────────────────────────────────────────
_reg_no_batch_index: Dict[str, str] = {}  # reg_no.lower() → sanitized_key

def _rebuild_reg_no_index() -> None:
    """Rebuild the full reg_no → batch_key index from all batch DB files."""
    global _reg_no_batch_index
    idx: Dict[str, str] = {}
    if not os.path.exists(BATCHES_DIR):
        _reg_no_batch_index = idx
        return
    for fname in os.listdir(BATCHES_DIR):
        if not (fname.startswith("batch_") and fname.endswith(".db")):
            continue
        raw_key = fname[len("batch_"):-len(".db")]
        db_path = os.path.join(BATCHES_DIR, fname)
        try:
            conn = sqlite3.connect(db_path)
            for (rn,) in conn.execute("SELECT LOWER(reg_no) FROM students WHERE reg_no IS NOT NULL").fetchall():
                if rn:
                    idx[rn] = raw_key
            conn.close()
        except Exception:
            pass
    _reg_no_batch_index = idx

def _index_add_student(reg_no: str, sanitized_key: str) -> None:
    """Add/update a single entry in the reg_no index."""
    if reg_no:
        _reg_no_batch_index[reg_no.strip().lower()] = sanitized_key

def _index_remove_student(reg_no: str) -> None:
    """Remove an entry from the reg_no index."""
    _reg_no_batch_index.pop((reg_no or "").strip().lower(), None)

def find_student_batch(reg_no: str) -> Optional[Tuple[str, str]]:
    """Find which batch DB holds a student by registration number.
    Returns (batch_name, sanitized_key) or None.

    Uses an O(1) in-memory index; rebuilds from disk on first call or cache miss.
    """
    clean_reg = (reg_no or "").strip().lower()
    if not clean_reg:
        return None

    # Lazy-build index on first use
    if not _reg_no_batch_index:
        _rebuild_reg_no_index()

    cached_key = _reg_no_batch_index.get(clean_reg)
    if cached_key:
        db_path = os.path.join(BATCHES_DIR, f"batch_{cached_key}.db")
        if os.path.exists(db_path):
            try:
                batch_db = get_batch_session(cached_key)
                stu = batch_db.query(Student).filter(func.lower(Student.reg_no) == clean_reg).first()
                batch_db.close()
                if stu:
                    return stu.batch or cached_key, cached_key
            except Exception:
                pass
        # Index is stale — fall through to full scan
        _index_remove_student(reg_no)

    # Full scan fallback (also refreshes the index for this student)
    if not os.path.exists(BATCHES_DIR):
        return None
    for fname in os.listdir(BATCHES_DIR):
        if not (fname.startswith("batch_") and fname.endswith(".db")):
            continue
        raw_key = fname[len("batch_"):-len(".db")]
        try:
            get_batch_engine(raw_key)  # ensure engine cached
            with _batch_sessionmakers[raw_key]() as s:
                stu = s.query(Student).filter(func.lower(Student.reg_no) == clean_reg).first()
                if stu:
                    _index_add_student(reg_no, raw_key)
                    return stu.batch or raw_key, raw_key
        except Exception:
            continue
    return None
    """Completely and permanently deletes all student and result data for a batch database."""
    sanitized = sanitize_batch_name(batch)
    if sanitized in _batch_engines:
        eng = _batch_engines.pop(sanitized, None)
        _batch_sessionmakers.pop(sanitized, None)
        if eng:
            try:
                eng.dispose()
            except Exception:
                pass

    candidates = {
        f"batch_{sanitized}.db",
        f"batch_{batch}.db",
        f"batch_{batch.replace('-', '_')}.db",
        f"batch_{batch.replace('_', '-')}.db"
    }
    deleted = False
    if os.path.exists(BATCHES_DIR):
        for c in candidates:
            p = os.path.join(BATCHES_DIR, c)
            if os.path.exists(p):
                try:
                    conn = sqlite3.connect(p)
                    conn.execute("DELETE FROM results")
                    conn.execute("DELETE FROM students")
                    conn.commit()
                    conn.execute("VACUUM")
                    conn.close()
                except Exception:
                    pass
                try:
                    os.remove(p)
                    deleted = True
                except Exception as e:
                    logger.warning("Could not delete file %s: %s", p, e)
                    deleted = True
            for ext in ("-journal", "-wal", "-shm"):
                jp = p + ext
                if os.path.exists(jp):
                    try:
                        os.remove(jp)
                    except Exception:
                        pass
    return deleted

def find_student_batch(reg_no: str) -> Optional[Tuple[str, str]]:
    """Find which batch and DB holds a student by registration number.
    Returns (batch_name, sanitized_key) or None."""
    clean_reg = (reg_no or "").strip().lower()
    if not clean_reg or not os.path.exists(BATCHES_DIR):
        return None
    for fname in os.listdir(BATCHES_DIR):
        if fname.startswith("batch_") and fname.endswith(".db"):
            db_path = os.path.join(BATCHES_DIR, fname)
            try:
                eng = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
                with sessionmaker(bind=eng)() as s:
                    stu = s.query(Student).filter(func.lower(Student.reg_no) == clean_reg).first()
                    if stu:
                        raw_key = fname[len("batch_"):-len(".db")]
                        return stu.batch or raw_key, raw_key
            except Exception:
                continue
    return None

def get_all_batch_sessions() -> List[Tuple[str, Session]]:
    """Yield open sessions for all existing batch databases."""
    sessions = []
    if os.path.exists(BATCHES_DIR):
        for fname in os.listdir(BATCHES_DIR):
            if fname.startswith("batch_") and fname.endswith(".db"):
                raw_key = fname[len("batch_"):-len(".db")]
                s = get_batch_session(raw_key)
                sessions.append((raw_key, s))
    return sessions

def get_subjects_map(db: Session) -> Dict[int, Subject]:
    """Cached lookup map of Subject.id -> Subject object."""
    subjects = db.query(Subject).all()
    return {sub.id: sub for sub in subjects}

# ==========================================
# PYDANTIC SCHEMAS
# ==========================================
class StudentBase(BaseModel):
    reg_no: str
    name: str
    department: str
    programme: Optional[str] = None
    batch: Optional[str] = None
    section: Optional[str] = None
    is_repeater: bool = False

class StudentCreate(StudentBase): pass
class StudentResponse(StudentBase):
    id: int
    email: Optional[str] = None
    model_config = ConfigDict(from_attributes=True)

class SubjectBase(BaseModel):
    code: str
    name: str
    credits: float = Field(gt=0)
    semester: str
    department: Optional[str] = None

class SubjectCreate(SubjectBase): pass
class SubjectResponse(SubjectBase):
    id: int
    model_config = ConfigDict(from_attributes=True)

class ResultBase(BaseModel):
    student_id: int
    subject_id: int
    semester: str
    year: Optional[str] = None
    grade: str
    batch: Optional[str] = None
    attempt: int = 1

class ResultCreate(ResultBase): pass

class ResultUpdate(BaseModel):
    grade: str
    year: Optional[str] = None
    batch: Optional[str] = None
    attempt: Optional[int] = None

class ResultResponse(ResultBase):
    id: int
    grade_point: float
    attempt: int = 1
    had_arrear: bool = False
    model_config = ConfigDict(from_attributes=True)

class ResultWithDetails(ResultResponse):
    student_name: str
    reg_no: str
    department: str
    subject_code: str
    subject_name: str
    credits: float

class GradeSummary(BaseModel):
    student_id: int
    reg_no: str
    name: str
    department: str
    semester: Optional[str] = None
    sgpa: Optional[float] = None
    cgpa: Optional[float] = None
    total_credits: float
    earned_credits: float
    grade_points_sum: float
    arrear_count: int = 0

class UploadResponse(BaseModel):
    message: str
    students_added: int
    results_added: int
    errors: List[str] = []

class OTPRequestRequest(BaseModel):
    reg_no: str
    email: str

class OTPRequestResponse(BaseModel):
    message: str
    expires_in_seconds: int
    resend_after_seconds: int

class OTPVerifyRequest(BaseModel):
    reg_no: str
    email: str
    otp: str

class OTPVerifyResponse(BaseModel):
    access_token: str
    expires_in_seconds: int

class ReportCardRequest(BaseModel):
    reg_no: str
    email: str
    access_token: str

class AttemptDetail(BaseModel):
    attempt: int
    semester: str
    year: Optional[str] = None
    grade: str
    grade_point: float
    is_cleared: bool = False

class ReportCardSubject(BaseModel):
    code: str
    name: str
    credits: float
    grade: str
    grade_point: float
    attempt: int = 1
    total_attempts: int = 1
    is_arrear: bool = False
    is_cleared: bool = False
    is_failed: bool = False
    failed_in_semester: Optional[str] = None
    cleared_in_semester: Optional[str] = None
    cleared_grade: Optional[str] = None
    original_semester: Optional[str] = None
    attempts_history: List[AttemptDetail] = []

class ArrearSubjectHistory(BaseModel):
    subject_code: str
    subject_name: str
    credits: float
    failed_semester: str
    failed_grade: str
    cleared_semester: Optional[str] = None
    cleared_grade: Optional[str] = None
    total_attempts: int = 1
    is_cleared: bool = False
    status: str

class ReportCardSemester(BaseModel):
    semester: str
    subjects: List[ReportCardSubject]
    sgpa: Optional[float] = None
    total_credits: float
    earned_credits: float

class ReportCardResponse(BaseModel):
    reg_no: str
    name: str
    department: str
    programme: Optional[str] = None
    batch: Optional[str] = None
    section: Optional[str] = None
    semesters: List[ReportCardSemester]
    arrear_history: List[ArrearSubjectHistory] = []
    has_arrears: bool = False
    cgpa: Optional[float] = None
    cgpa_percentage: Optional[float] = None
    total_credits: float
    earned_credits: float

# ==========================================
# AUTH DEPENDENCIES
# ==========================================
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode('utf-8'), hashed.encode('utf-8'))

def get_system_db():
    db = SystemSessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_subjects_db():
    db = SubjectsSessionLocal()
    try:
        yield db
    finally:
        db.close()

# Alias get_db for routes that primarily need system/auth session
get_db = get_system_db

def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_system_db)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if not username:
            raise HTTPException(status_code=401, detail="Invalid token")
        user = db.query(User).filter(User.username == username).first()
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

ACCOUNT_TYPES = ["Admin", "TNP", "Faculty", "Exam Wing"]

def validate_password_strength(password: str):
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters long.")
    if not re.search(r"[a-z]", password):
        raise HTTPException(status_code=400, detail="Password must contain at least one lowercase letter.")
    if not re.search(r"[A-Z]", password):
        raise HTTPException(status_code=400, detail="Password must contain at least one uppercase letter.")
    if not re.search(r"\d", password):
        raise HTTPException(status_code=400, detail="Password must contain at least one digit.")
    if not re.search(r"[!@#$%^&*]", password):
        raise HTTPException(status_code=400, detail="Password must contain a special character (!@#$%^&*).")

def role_checker(allowed_roles: List[str]):
    allowed_lower = [r.lower() for r in allowed_roles]
    def checker(user: User = Depends(get_current_user)):
        if not user.role or user.role.lower() not in allowed_lower:
            raise HTTPException(status_code=403, detail="Insufficient privileges")
        return user
    return checker

allow_all = role_checker(["Admin", "TNP", "Faculty", "Exam Wing", "manager", "viewer"])
allow_write = role_checker(["Admin", "Exam Wing", "manager"])
allow_admin = role_checker(["Admin"])

# ==========================================
# REPORT CARD OTP HELPERS
# ==========================================
def _generate_otp() -> str:
    return f"{secrets.randbelow(10 ** OTP_LENGTH):0{OTP_LENGTH}d}"

def _hash_otp(otp: str) -> str:
    return bcrypt.hashpw(otp.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def _verify_otp_hash(otp: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(otp.encode("utf-8"), hashed.encode("utf-8"))
    except ValueError:
        return False

def _send_via_resend(api_key: str, from_email: str, from_name: str, to_email: str, subject: str, text_body: str, html_body: str) -> bool:
    url = "https://api.resend.com/emails"
    custom_from = os.getenv("RESEND_FROM_EMAIL", "").strip()
    if custom_from:
        sender = f"{from_name} <{custom_from}>" if from_name and "<" not in custom_from else custom_from
    elif from_email and not from_email.lower().endswith(("@gmail.com", "@googlemail.com", "@yahoo.com", "@outlook.com", "@hotmail.com")):
        sender = f"{from_name} <{from_email}>" if from_name else from_email
    else:
        sender = f"{from_name} <onboarding@resend.dev>" if from_name else "onboarding@resend.dev"

    payload = json.dumps({
        "from": sender,
        "to": [to_email],
        "subject": subject,
        "html": html_body,
        "text": text_body
    }).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=payload,
        headers={
            "Authorization": f"Bearer {api_key.strip()}",
            "Content-Type": "application/json",
            "User-Agent": "PTU-Grade-Portal/1.0"
        }
    )
    try:
        with urllib.request.urlopen(req, timeout=8) as resp:
            return 200 <= resp.status < 300
    except urllib.error.HTTPError as e:
        err_msg = e.read().decode("utf-8", errors="ignore")
        logger.error("Resend API HTTP %s: %s (from=%s, to=%s)", e.code, err_msg, sender, to_email)
        return False
    except Exception as e:
        logger.error("Resend API failed: %s", e)
        return False


def _send_via_brevo(api_key: str, from_email: str, from_name: str, to_email: str, subject: str, text_body: str, html_body: str) -> bool:
    url = "https://api.brevo.com/v3/smtp/email"
    sender_email = os.getenv("BREVO_FROM_EMAIL", "").strip() or from_email or "velshakthi406@gmail.com"
    payload = json.dumps({
        "sender": {"name": from_name or "PTU Grade Portal", "email": sender_email},
        "to": [{"email": to_email}],
        "subject": subject,
        "htmlContent": html_body,
        "textContent": text_body
    }).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=payload,
        headers={
            "api-key": api_key.strip(),
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": "PTU-Grade-Portal/1.0"
        }
    )
    try:
        with urllib.request.urlopen(req, timeout=8) as resp:
            return 200 <= resp.status < 300
    except urllib.error.HTTPError as e:
        err_msg = e.read().decode("utf-8", errors="ignore")
        logger.error("Brevo API HTTP %s: %s (sender=%s, to=%s)", e.code, err_msg, sender_email, to_email)
        return False
    except Exception as e:
        logger.error("Brevo API failed: %s", e)
        return False


def _send_via_sendgrid(api_key: str, from_email: str, from_name: str, to_email: str, subject: str, text_body: str, html_body: str) -> bool:
    url = "https://api.sendgrid.com/v3/mail/send"
    payload = json.dumps({
        "personalizations": [{"to": [{"email": to_email}]}],
        "from": {"email": from_email, "name": from_name or "PTU Grade Portal"},
        "subject": subject,
        "content": [
            {"type": "text/plain", "value": text_body},
            {"type": "text/html", "value": html_body}
        ]
    }).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=payload,
        headers={
            "Authorization": f"Bearer {api_key.strip()}",
            "Content-Type": "application/json",
            "User-Agent": "PTU-Grade-Portal/1.0"
        }
    )
    try:
        with urllib.request.urlopen(req, timeout=8) as resp:
            return 200 <= resp.status < 300
    except urllib.error.HTTPError as e:
        err_msg = e.read().decode("utf-8", errors="ignore")
        logger.error("SendGrid API HTTP %s: %s", e.code, err_msg)
        return False
    except Exception as e:
        logger.error("SendGrid API failed: %s", e)
        return False


def _send_via_smtp(to_email: str, subject: str, text_body: str, html_body: str) -> bool:
    message = MIMEMultipart("alternative")
    message["Subject"] = subject
    message["From"] = f"{SMTP_FROM_NAME} <{SMTP_FROM_EMAIL}>"
    message["To"] = to_email
    message.attach(MIMEText(text_body, "plain"))
    message.attach(MIMEText(html_body, "html"))

    if SMTP_USE_SSL:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=context, timeout=5) as server:
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(SMTP_FROM_EMAIL, [to_email], message.as_string())
    else:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=5) as server:
            server.ehlo()
            server.starttls(context=ssl.create_default_context())
            server.ehlo()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(SMTP_FROM_EMAIL, [to_email], message.as_string())
    return True


def _dispatch_email(to_email: str, subject: str, text_body: str, html_body: str) -> bool:
    """Send email using HTTPS APIs (works on Render free tier where SMTP is blocked) or standard SMTP."""
    # 1. Brevo API (HTTP port 443 - works without domain verification)
    if BREVO_API_KEY:
        if _send_via_brevo(BREVO_API_KEY, SMTP_FROM_EMAIL, SMTP_FROM_NAME, to_email, subject, text_body, html_body):
            return True

    # 2. Resend API (HTTP port 443 - requires verified custom domain for outside recipients)
    if RESEND_API_KEY:
        if _send_via_resend(RESEND_API_KEY, SMTP_FROM_EMAIL, SMTP_FROM_NAME, to_email, subject, text_body, html_body):
            return True

    # 3. SendGrid API (HTTP port 443)
    if SENDGRID_API_KEY:
        if _send_via_sendgrid(SENDGRID_API_KEY, SMTP_FROM_EMAIL, SMTP_FROM_NAME, to_email, subject, text_body, html_body):
            return True

    # 4. Standard SMTP (works on local machine, VPS, or paid cloud instances)
    if SMTP_CONFIGURED:
        try:
            return _send_via_smtp(to_email, subject, text_body, html_body)
        except Exception as e:
            logger.warning("SMTP delivery failed (expected on Render free tier where ports 25/465/587 are blocked): %s", e)
            return False

    return False


def _send_otp_email(to_email: str, student_name: str, reg_no: str, otp: str) -> bool:
    subject = "Your Report Card OTP - PTU Grade Portal"
    text_body = (
        f"Hello {student_name},\n\n"
        f"Your one-time passcode (OTP) to view your report card is:\n\n"
        f"    {otp}\n\n"
        f"This code is valid for {OTP_EXPIRE_MINUTES} minutes and can be used only once.\n"
        f"Register Number: {reg_no}\n\n"
        f"If you did not request this, you can safely ignore this email.\n\n"
        f"- PTU Grade Portal"
    )
    html_body = f"""
    <div style="font-family:Segoe UI,Arial,sans-serif;max-width:480px;margin:auto;
                border:1px solid #e5e5e5;border-radius:12px;overflow:hidden;">
      <div style="background:#7a1230;color:#fff;padding:20px 24px;">
        <h2 style="margin:0;font-size:18px;">PTU Grade Portal</h2>
      </div>
      <div style="padding:24px;">
        <p>Hello <strong>{student_name}</strong> ({reg_no}),</p>
        <p>Use the code below to verify your identity and view your report card:</p>
        <div style="text-align:center;margin:24px 0;">
          <span style="display:inline-block;font-size:32px;letter-spacing:8px;
                       font-weight:700;color:#7a1230;background:#f6f0f2;
                       padding:12px 20px;border-radius:8px;">{otp}</span>
        </div>
        <p style="color:#555;font-size:14px;">
          This code expires in {OTP_EXPIRE_MINUTES} minutes and can only be used once.
          Never share it with anyone.
        </p>
        <p style="color:#999;font-size:12px;">
          If you did not request this, you can safely ignore this email.
        </p>
      </div>
    </div>
    """

    sent = _dispatch_email(to_email, subject, text_body, html_body)
    if not sent:
        logger.error(
            "Failed to deliver Report Card OTP email to %s <%s>. (Check email service credentials, sender verification, or provider quotas)",
            reg_no, to_email
        )
    return sent

STAFF_OTP_EXPIRE_MINUTES = 10
STAFF_OTP_MAX_ATTEMPTS = 5
STAFF_OTP_TOKEN_SECRET = os.getenv("STAFF_OTP_TOKEN_SECRET", SECRET_KEY + "_staff_otp")
STAFF_OTP_TOKEN_EXPIRE_MINUTES = 30   # window to set password after OTP verify

def _send_staff_otp_email(to_email: str, name: str, otp: str, purpose: str) -> bool:
    """Send an OTP email to a staff member for registration or password reset."""
    action = "Registration Verification" if purpose == "register" else "Password Reset"
    subject = f"Your {action} OTP — PTU Grade Portal"
    text_body = (
        f"Hello {name},\n\n"
        f"Your one-time passcode (OTP) for {action.lower()} is:\n\n"
        f"    {otp}\n\n"
        f"This code is valid for {STAFF_OTP_EXPIRE_MINUTES} minutes and can be used only once.\n\n"
        f"If you did not request this, please ignore this email.\n\n"
        f"— PTU Grade Portal"
    )
    html_body = f"""
    <div style="font-family:Segoe UI,Arial,sans-serif;max-width:480px;margin:auto;
                border:1px solid #e5e5e5;border-radius:12px;overflow:hidden;">
      <div style="background:#7a1230;color:#fff;padding:20px 24px;">
        <h2 style="margin:0;font-size:18px;">PTU Grade Portal – Staff Portal</h2>
      </div>
      <div style="padding:24px;">
        <p>Hello <strong>{name}</strong>,</p>
        <p>Your one-time passcode for <strong>{action}</strong>:</p>
        <div style="text-align:center;margin:24px 0;">
          <span style="display:inline-block;font-size:32px;letter-spacing:8px;
                       font-weight:700;color:#7a1230;background:#f6f0f2;
                       padding:12px 20px;border-radius:8px;">{otp}</span>
        </div>
        <p style="color:#555;font-size:14px;">
          This code expires in {STAFF_OTP_EXPIRE_MINUTES} minutes and can only be used once.
          Never share it with anyone.
        </p>
        <p style="color:#999;font-size:12px;">
          If you did not request this, please ignore this email.
        </p>
      </div>
    </div>
    """

    sent = _dispatch_email(to_email, subject, text_body, html_body)
    if not sent:
        logger.error(
            "Failed to deliver Staff OTP email to %s <%s> purpose=%s.",
            name, to_email, purpose
        )
    return sent

def _issue_staff_otp_token(email: str, purpose: str) -> str:
    payload = {
        "purpose": f"staff_{purpose}",
        "email": email,
        "exp": datetime.datetime.utcnow() + datetime.timedelta(minutes=STAFF_OTP_TOKEN_EXPIRE_MINUTES),
    }
    return jwt.encode(payload, STAFF_OTP_TOKEN_SECRET, algorithm=ALGORITHM)

def _verify_staff_otp_token(token: str, email: str, purpose: str) -> None:
    try:
        payload = jwt.decode(token, STAFF_OTP_TOKEN_SECRET, algorithms=[ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Your verification session has expired. Please request a new OTP.")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid verification session. Please request a new OTP.")

    if (
        payload.get("purpose") != f"staff_{purpose}"
        or payload.get("email", "").strip().lower() != email.strip().lower()
    ):
        raise HTTPException(status_code=401, detail="Invalid verification session. Please request a new OTP.")


def _issue_report_card_token(reg_no: str, email: str) -> str:
    payload = {
        "purpose": "report_card_access",
        "reg_no": reg_no,
        "email": email,
        "exp": datetime.datetime.utcnow() + datetime.timedelta(minutes=REPORT_CARD_TOKEN_EXPIRE_MINUTES),
    }
    return jwt.encode(payload, REPORT_CARD_TOKEN_SECRET, algorithm=ALGORITHM)

def _verify_report_card_token(token: str, reg_no: str, email: str) -> None:
    try:
        payload = jwt.decode(token, REPORT_CARD_TOKEN_SECRET, algorithms=[ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Your verification session has expired. Please verify the OTP again.")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid verification session. Please verify the OTP again.")

    if (
        payload.get("purpose") != "report_card_access"
        or payload.get("reg_no", "").strip().lower() != reg_no.strip().lower()
        or payload.get("email", "").strip().lower() != email.strip().lower()
    ):
        raise HTTPException(status_code=401, detail="Invalid verification session. Please verify the OTP again.")

# ==========================================
# FASTAPI APP
# ==========================================
app = FastAPI(title="University Grade Processing System")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# ==========================================
# AUTH ROUTES
# ==========================================

# ── OTP Request ──────────────────────────────
class StaffOtpRequestBody(BaseModel):
    name: str              # Full name (required for 'register', optional hint for 'reset')
    email: str
    purpose: str           # 'register' or 'reset'

@app.post("/auth/otp/request")
def request_staff_otp(data: StaffOtpRequestBody, db: Session = Depends(get_system_db)):
    if data.purpose not in ("register", "reset"):
        raise HTTPException(status_code=400, detail="purpose must be 'register' or 'reset'")

    clean_email = str(data.email).strip().lower()
    clean_name  = data.name.strip()

    if data.purpose == "register":
        # Must pre-exist in Resources with matching name + email
        resource = db.query(Resource).filter(
            func.lower(Resource.name) == clean_name.lower(),
            func.lower(Resource.email) == clean_email
        ).first()
        if not resource:
            raise HTTPException(
                status_code=403,
                detail="No resource record found with this name and email. Please contact your administrator."
            )
        # Must not already have an account
        existing = db.query(User).filter(
            func.lower(User.email) == clean_email
        ).first()
        if existing:
            raise HTTPException(
                status_code=400,
                detail="This email already has a registered account. Please sign in."
            )
        display_name = clean_name

    else:  # reset
        # Must have an existing account
        user = db.query(User).filter(
            func.lower(User.email) == clean_email
        ).first()
        if not user:
            raise HTTPException(
                status_code=404,
                detail="No account found with this email address."
            )
        display_name = user.username

    # Invalidate any previous unverified OTPs for this email+purpose
    db.query(StaffOTP).filter(
        func.lower(StaffOTP.email) == clean_email,
        StaffOTP.purpose == data.purpose,
        StaffOTP.verified == False
    ).delete(synchronize_session=False)
    db.commit()

    otp = _generate_otp()
    otp_hash = _hash_otp(otp)
    expires_at = datetime.datetime.utcnow() + datetime.timedelta(minutes=STAFF_OTP_EXPIRE_MINUTES)

    staff_otp = StaffOTP(
        email=clean_email,
        purpose=data.purpose,
        otp_hash=otp_hash,
        verified=False,
        attempts=0,
        expires_at=expires_at,
    )
    db.add(staff_otp)
    db.commit()

    sent = _send_staff_otp_email(to_email=clean_email, name=display_name, otp=otp, purpose=data.purpose)
    if not sent:
        db.delete(staff_otp)
        db.commit()
        raise HTTPException(
            status_code=503,
            detail="Unable to deliver OTP email. Please check your email configuration on the server."
        )

    return {"message": f"OTP sent to {clean_email}. Valid for {STAFF_OTP_EXPIRE_MINUTES} minutes."}


# ── OTP Verify ───────────────────────────────
class StaffOtpVerifyBody(BaseModel):
    email: str
    otp: str
    purpose: str

@app.post("/auth/otp/verify")
def verify_staff_otp(data: StaffOtpVerifyBody, db: Session = Depends(get_system_db)):
    if data.purpose not in ("register", "reset"):
        raise HTTPException(status_code=400, detail="purpose must be 'register' or 'reset'")

    clean_email = str(data.email).strip().lower()

    record = db.query(StaffOTP).filter(
        func.lower(StaffOTP.email) == clean_email,
        StaffOTP.purpose == data.purpose,
        StaffOTP.verified == False
    ).order_by(StaffOTP.created_at.desc()).first()

    if not record:
        raise HTTPException(status_code=400, detail="No active OTP found. Please request a new one.")

    if datetime.datetime.utcnow() > record.expires_at:
        db.delete(record)
        db.commit()
        raise HTTPException(status_code=400, detail="OTP has expired. Please request a new one.")

    if record.attempts >= STAFF_OTP_MAX_ATTEMPTS:
        db.delete(record)
        db.commit()
        raise HTTPException(status_code=400, detail="Too many failed attempts. Please request a new OTP.")

    record.attempts += 1
    db.commit()

    if not _verify_otp_hash(data.otp.strip(), record.otp_hash):
        remaining = STAFF_OTP_MAX_ATTEMPTS - record.attempts
        raise HTTPException(
            status_code=400,
            detail=f"Incorrect OTP. {remaining} attempt(s) remaining."
        )

    # Mark as verified
    record.verified = True
    db.commit()

    # Issue a short-lived token to allow the password step
    otp_token = _issue_staff_otp_token(email=clean_email, purpose=data.purpose)
    return {"otp_token": otp_token, "message": "OTP verified successfully. You may now set your password."}


# ── Register (complete with password) ────────
class StaffRegisterBody(BaseModel):
    name: str
    email: str
    otp_token: str
    password: str
    confirm_password: str

@app.post("/auth/register")
def register_staff(data: StaffRegisterBody, db: Session = Depends(get_system_db)):
    clean_email = str(data.email).strip().lower()
    clean_name  = data.name.strip()

    # Validate OTP token
    _verify_staff_otp_token(data.otp_token, email=clean_email, purpose="register")

    # Re-verify resource exists
    resource = db.query(Resource).filter(
        func.lower(Resource.name) == clean_name.lower(),
        func.lower(Resource.email) == clean_email
    ).first()
    if not resource:
        raise HTTPException(
            status_code=403,
            detail="Resource record not found. Please contact your administrator."
        )

    # Ensure no account exists
    existing = db.query(User).filter(
        func.lower(User.email) == clean_email
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="An account with this email already exists. Please sign in.")

    # Validate passwords
    if data.password != data.confirm_password:
        raise HTTPException(status_code=400, detail="Passwords do not match.")
    validate_password_strength(data.password)

    # Derive a unique username from the email local part (before the @)
    # This ensures uniqueness since email is already unique.
    base_username = clean_email.split("@")[0]

    # Create user with role from resource
    hashed = hash_password(data.password)
    user = User(
        username=base_username,
        email=clean_email,
        hashed_password=hashed,
        role=resource.account_type
    )
    db.add(user)

    # Clean up verified OTP records for this email
    db.query(StaffOTP).filter(
        func.lower(StaffOTP.email) == clean_email,
        StaffOTP.purpose == "register"
    ).delete(synchronize_session=False)

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="An account with this email or username already exists. Please sign in.")

    return {"message": "Registration complete! You can now sign in.", "role": user.role, "username": user.username}


# ── Reset Password ────────────────────────────
class StaffResetPasswordBody(BaseModel):
    email: str
    otp_token: str
    new_password: str
    confirm_password: str

@app.post("/auth/reset-password")
def reset_staff_password(data: StaffResetPasswordBody, db: Session = Depends(get_system_db)):
    clean_email = str(data.email).strip().lower()

    # Validate OTP token
    _verify_staff_otp_token(data.otp_token, email=clean_email, purpose="reset")

    user = db.query(User).filter(func.lower(User.email) == clean_email).first()
    if not user:
        raise HTTPException(status_code=404, detail="Account not found.")

    if data.new_password != data.confirm_password:
        raise HTTPException(status_code=400, detail="Passwords do not match.")
    validate_password_strength(data.new_password)

    user.hashed_password = hash_password(data.new_password)

    # Clean up reset OTP records
    db.query(StaffOTP).filter(
        func.lower(StaffOTP.email) == clean_email,
        StaffOTP.purpose == "reset"
    ).delete(synchronize_session=False)

    db.commit()
    return {"message": "Password reset successfully. You can now sign in with your new password."}


@app.post("/auth/login")
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_system_db)):
    ident = form_data.username.strip().lower()
    user = db.query(User).filter(
        (func.lower(User.username) == ident) |
        (func.lower(User.email) == ident)
    ).first()

    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(status_code=400, detail="Invalid credentials")

    # Verify that the matching resource record pre-exists!
    resource = db.query(Resource).filter(
        (func.lower(Resource.email) == func.lower(user.email or "")) |
        (func.lower(Resource.name) == func.lower(user.username))
    ).first()
    if not resource:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Your associated resource record no longer exists. Access denied."
        )

    # Sync role with resource account_type
    if user.role != resource.account_type:
        user.role = resource.account_type
        db.commit()

    access_token = jwt.encode(
        {"sub": user.username, "role": user.role, "exp": datetime.datetime.utcnow() + datetime.timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)},
        SECRET_KEY, algorithm=ALGORITHM
    )
    refresh_token = jwt.encode(
        {"sub": user.username, "role": user.role, "exp": datetime.datetime.utcnow() + datetime.timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)},
        REFRESH_SECRET_KEY, algorithm=ALGORITHM
    )
    return {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "token_type": "bearer",
        "role": user.role,
        "username": user.username
    }

@app.get("/auth/me")
def get_me(current_user: User = Depends(get_current_user)):
    return {
        "id": current_user.id,
        "username": current_user.username,
        "email": current_user.email,
        "role": current_user.role
    }

@app.get("/auth/allowed-types")
def get_allowed_account_types(current_user: User = Depends(allow_all)):
    """Return account types the current user is allowed to assign when creating resources."""
    role = (current_user.role or "").strip()
    if role.lower() == "admin":
        return {"allowed": ACCOUNT_TYPES}
    # Non-admin: can only create resources of their own type
    if role in ACCOUNT_TYPES:
        return {"allowed": [role]}
    return {"allowed": []}



# ==========================================
# RESOURCE SCHEMAS & ROUTES
# ==========================================
class ResourceBase(BaseModel):
    name: str = Field(..., min_length=1)
    email: str
    account_type: str

    @field_validator('account_type')
    @classmethod
    def validate_account_type(cls, v: str) -> str:
        for t in ACCOUNT_TYPES:
            if v.strip().lower() == t.lower():
                return t
        raise ValueError(f"Account type must be one of: {', '.join(ACCOUNT_TYPES)}")

class ResourceCreate(ResourceBase):
    pass

class ResourceUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    account_type: Optional[str] = None

class ResourceResponse(ResourceBase):
    id: int
    has_account: bool = False
    created_at: Optional[datetime.datetime] = None
    model_config = ConfigDict(from_attributes=True)

@app.get("/resources/", response_model=List[ResourceResponse])
def get_resources(db: Session = Depends(get_system_db), user: User = Depends(allow_all)):
    resources = db.query(Resource).order_by(Resource.id.asc()).all()
    user_names_and_emails = set()
    for u in db.query(User.username, User.email).all():
        if u[0]:
            user_names_and_emails.add(u[0].strip().lower())
        if u[1]:
            user_names_and_emails.add(u[1].strip().lower())

    result = []
    for r in resources:
        has_acc = (r.name.strip().lower() in user_names_and_emails or r.email.strip().lower() in user_names_and_emails)
        result.append(ResourceResponse(
            id=r.id,
            name=r.name,
            email=r.email,
            account_type=r.account_type,
            has_account=has_acc,
            created_at=r.created_at
        ))
    return result

@app.post("/resources/", response_model=ResourceResponse)
def create_resource(data: ResourceCreate, db: Session = Depends(get_system_db), user: User = Depends(allow_all)):
    clean_email = str(data.email).strip().lower()
    clean_name = data.name.strip()
    requester_role = (user.role or "").strip()

    # --- Role-based type enforcement ---
    if requester_role.lower() != "admin":
        # Non-admin: must only create resources of their own account type
        if data.account_type.lower() != requester_role.lower():
            raise HTTPException(
                status_code=403,
                detail=f"You can only create resources of your own account type ({requester_role})."
            )

    # --- Email must be unique (name can repeat) ---
    existing_email = db.query(Resource).filter(
        func.lower(Resource.email) == clean_email
    ).first()
    if existing_email:
        raise HTTPException(status_code=400, detail="A resource with this email already exists.")

    new_res = Resource(
        name=clean_name,
        email=clean_email,
        account_type=data.account_type
    )
    db.add(new_res)
    db.commit()
    db.refresh(new_res)
    return ResourceResponse(
        id=new_res.id,
        name=new_res.name,
        email=new_res.email,
        account_type=new_res.account_type,
        has_account=False,
        created_at=new_res.created_at
    )

@app.put("/resources/{resource_id}", response_model=ResourceResponse)
def update_resource(resource_id: int, data: ResourceUpdate, db: Session = Depends(get_system_db), user: User = Depends(allow_admin)):
    res = db.query(Resource).filter(Resource.id == resource_id).first()
    if not res:
        raise HTTPException(status_code=404, detail="Resource not found.")

    old_email = res.email

    # Name can now repeat — just update it without a uniqueness check
    if data.name is not None and data.name.strip():
        res.name = data.name.strip()

    if data.email is not None and str(data.email).strip():
        clean_email = str(data.email).strip().lower()
        dup_email = db.query(Resource).filter(
            func.lower(Resource.email) == clean_email,
            Resource.id != resource_id
        ).first()
        if dup_email:
            raise HTTPException(status_code=400, detail="Another resource already uses this email.")
        res.email = clean_email

    if data.account_type is not None and data.account_type.strip():
        # Admin cannot change their OWN account type
        own_res = db.query(Resource).filter(
            func.lower(Resource.email) == func.lower(user.email or ""),
            Resource.id == resource_id
        ).first()
        if own_res:
            raise HTTPException(
                status_code=403,
                detail="You cannot change your own account type."
            )

        # Remove Admin choice in change account type
        if data.account_type.strip().lower() == "admin":
            raise HTTPException(
                status_code=400,
                detail="Account type cannot be changed to Admin."
            )

        valid_type = None
        for t in ["TNP", "Faculty", "Exam Wing"]:
            if data.account_type.strip().lower() == t.lower():
                valid_type = t
                break
        if not valid_type:
            raise HTTPException(status_code=400, detail="Account type must be one of: TNP, Faculty, Exam Wing")
        res.account_type = valid_type

    # Sync active User record if one exists for this resource (match by old email)
    active_user = db.query(User).filter(
        func.lower(User.email) == old_email.lower()
    ).first()
    if active_user:
        active_user.role = res.account_type
        active_user.username = res.name
        active_user.email = res.email

    db.commit()
    db.refresh(res)

    has_acc = db.query(User).filter(
        func.lower(User.email) == res.email.lower()
    ).first() is not None

    return ResourceResponse(
        id=res.id,
        name=res.name,
        email=res.email,
        account_type=res.account_type,
        has_account=has_acc,
        created_at=res.created_at
    )

@app.delete("/resources/{resource_id}")
def delete_resource(resource_id: int, db: Session = Depends(get_system_db), user: User = Depends(allow_admin)):
    res = db.query(Resource).filter(Resource.id == resource_id).first()
    if not res:
        raise HTTPException(status_code=404, detail="Resource not found.")

    if res.email.lower() == (user.email or "").lower():
        raise HTTPException(status_code=400, detail="You cannot delete your own active resource record.")

    # Remove linked user account
    linked_user = db.query(User).filter(
        func.lower(User.email) == res.email.lower()
    ).first()
    if linked_user:
        db.delete(linked_user)

    db.delete(res)
    db.commit()
    return {"message": f"Resource '{res.name}' deleted successfully."}


@app.post("/auth/refresh")
def refresh(refresh_token: str):
    try:
        payload = jwt.decode(refresh_token, REFRESH_SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        role = payload.get("role")
        new_access = jwt.encode(
            {"sub": username, "role": role, "exp": datetime.datetime.utcnow() + datetime.timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)},
            SECRET_KEY, algorithm=ALGORITHM
        )
        new_refresh = jwt.encode(
            {"sub": username, "role": role, "exp": datetime.datetime.utcnow() + datetime.timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)},
            REFRESH_SECRET_KEY, algorithm=ALGORITHM
        )
        return {"access_token": new_access, "refresh_token": new_refresh, "token_type": "bearer"}
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

# ==========================================
# SUBJECT PARSER
# ==========================================
CODE_PATTERN = r'[A-Z]{2}U[A-Z]\d{2,4}[A-Za-z]?'
CCC_TOKENS = r'(?:BSC|PCC|ESC|AEC|SEC|VAC|ANC|PEC|HNC|ASC)\*?'
VALID_ROMANS = {"I", "II", "III", "IV", "V", "VI", "VII", "VIII"}

def _normalize_line(line: str) -> str:
    if not line:
        return line
    line = line.replace('\u2013', '-').replace('\u2014', '-')
    line = line.replace('\xa0', ' ')
    line = re.sub(r'[ \t]+', ' ', line)
    return line.strip()

DEPT_PREFIX_MAP = {
    "CE": "CE", "ME": "ME", "EC": "ECE", "CS": "CSE", "EE": "EEE",
    "EI": "EIE", "CH": "CHE", "IT": "IT", "MT": "MT", "MA": "Maths",
    "PH": "Physics", "CY": "Chemistry", "HS": "HS", "GE": "GE",
}

def infer_department_from_code(code: str) -> Optional[str]:
    m = re.match(r'^([A-Z]{2})U', code)
    if not m:
        return None
    return DEPT_PREFIX_MAP.get(m.group(1))

CCC_ONLY_RE = re.compile(rf'^{CCC_TOKENS}$')

def _clean_name(name: str) -> str:
    name = re.sub(r'\s+', ' ', name).strip()
    name = name.strip(' -–')
    words = name.split(' ')
    words = [w for w in words if not CCC_ONLY_RE.match(w)]
    deduped: List[str] = []
    for w in words:
        if deduped and deduped[-1].lower() == w.lower():
            continue
        deduped.append(w)
    name = ' '.join(deduped).strip()
    name = name.strip(' -–')
    return name

def _last_number(rest: str) -> Optional[float]:
    nums = re.findall(r'\d+(?:\.\d+)?', rest)
    if not nums:
        return None
    try:
        return float(nums[-1])
    except ValueError:
        return None

SUMMARY_LINE_RE = re.compile(
    rf'^(?P<code>{CODE_PATTERN})\s+(?P<name>.+?)\s+'
    rf'(?P<sem>I{{1,3}}V?|IV|VI{{0,3}}|VIII|VII|V)\s+(?P<credit>\d+(?:\.\d+)?)\s*$'
)

MAIN_LINE_RE = re.compile(
    rf'^(?P<code>{CODE_PATTERN})\s+(?P<name>.+?)\s+(?P<ccc>{CCC_TOKENS})\s+(?P<rest>[\d.\-_\s]+)$'
)

ORPHAN_LINE_RE = re.compile(
    rf'^(?P<code>{CODE_PATTERN})\s+(?:(?P<ccc>{CCC_TOKENS})\s+)?(?P<rest>[\d.\-_\s]+)$'
)

ORPHAN_SUMMARY_LINE_RE = re.compile(
    rf'^(?P<code>{CODE_PATTERN})\s+'
    rf'(?:(?P<sem>I{{1,3}}V?|IV|VI{{0,3}}|VIII|VII|V)\s+)?'
    rf'(?P<credit>\d+(?:\.\d+)?)\s*$'
)

SEM_HEADING_RE = re.compile(r'^Semester\s+([IVX]+)\s*$')

FRAGMENT_BLACKLIST_RE = re.compile(
    r'^(Total\b|List of|Course\s*Code|CCC\b|Ancillary [Ss]tream|Exit Option|'
    r'Exit option|Note[: ]|Offered to|Other Department Courses|Professional Elective\b|'
    r'For (Other|other)|Interdisciplinary|Courses? [Oo]ffered|COURSES? OFFERED|'
    r'LIST OF|XX\s*[–-]|xx-|Group-|Semester\b|3 [Ww]eeks|Induction Program|'
    r'\(For |\(Offered|\(Interdisciplinary|Periods\b|Credits\b|L\s+T\s+P\b|'
    r'^[LTP]\s*[#*]?$|XXX\b|Serial Number|'
    r'^Professional$|^Elective\b|^Course$|^Code$|Course\s+Semester|Elective\s+Code)',
    re.IGNORECASE
)

PLACEHOLDER_NAME_RE = re.compile(
    r'^Professional Elective\s*[-–]?\s*(I{1,3}|IV|VI{0,3}|VII|VIII|V|\d+)$',
    re.IGNORECASE
)

_ROMAN_RE = re.compile(r'^(I{1,3}|IV|VI{0,3}|VIII|VII|V|IX|X)$')
ROMAN_LIST_RE = re.compile(
    r'\b(I{1,3}|IV|VI{0,3}|VIII|VII|V)\b\s*,\s*\b(I{1,3}|IV|VI{0,3}|VIII|VII|V)\b'
)
PEC_LABEL_TO_SEM = {"I": "V", "II": "VI", "III": "VII"}
PEC_HEADER_PREFIX_RE = re.compile(
    r'^(?:Professional\s+)?Elective\s*[-–]?\s*(?P<label>I{1,3}|IV|VI{0,3}|VIII|VII|V)\b/?'
    r'(?:\s*/?\s*(?:Professional\s+)?Elective\s*[-–]?\s*(?:I{1,3}|IV|VI{0,3}|VIII|VII|V)\b/?)*'
    r'\s*',
    re.IGNORECASE
)
PEC_SECTION_START_RE = re.compile(
    r'^List of Professional Elective Courses|^LIST OF PROFESSIONAL ELECTIVES?\b',
    re.IGNORECASE
)
PEC_LINE_RE = re.compile(
    rf'^(?P<code>{CODE_PATTERN})\s+(?P<name>.+?)'
    rf'(?:\s+(?P<sem>I{{1,3}}|IV|VI{{0,3}}|VIII|VII|V)/?)?\s*$'
)
ANC_LINE_RE = re.compile(
    rf'^(?P<code>{CODE_PATTERN})\s+(?P<name>.+?)\s*$'
)
ANC_LINE_PREFIXED_RE = re.compile(
    rf'^(?:(?!{CODE_PATTERN}).)+?(?P<code>{CODE_PATTERN})\s+(?P<name>.+?)\s*$'
)

def _pec_position_sem(position: int) -> str:
    if position <= 5:
        return "V"
    if position <= 10:
        return "VI"
    return "VII"

ANC_SET_POSITION_TO_SEM = {1: "IV", 2: "V", 3: "VI", 4: "VII"}

def _anc_position_sem(position: int) -> str:
    slot = ((position - 1) % 4) + 1
    return ANC_SET_POSITION_TO_SEM[slot]

ANC_HEADING_RE = re.compile(r'\bAncillary\b', re.IGNORECASE)

def _is_plain_fragment(line: str) -> bool:
    if not line or FRAGMENT_BLACKLIST_RE.search(line) or re.search(CODE_PATTERN, line):
        return False
    if re.search(r'[A-Z]{2,6}XXX+\b', line):
        return False
    if re.search(r'\d', line):
        words = line.split()
        if all(re.fullmatch(r'[\d.\-_]+', w) or CCC_ONLY_RE.match(w) for w in words):
            return False
    if _ROMAN_RE.match(line) or ROMAN_LIST_RE.search(line) or CCC_ONLY_RE.match(line):
        return False
    words = line.split()
    if not words or len(words) > 8:
        return False
    return True

def parse_subjects_pdf(file_content: bytes) -> List[Dict[str, Any]]:
    with pdfplumber.open(io.BytesIO(file_content)) as pdf:
        all_lines: List[str] = []
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            all_lines.extend(_normalize_line(line) for line in text.split('\n'))

    if not all_lines:
        return []

    n = len(all_lines)
    subjects: Dict[str, Dict[str, Any]] = {}
    source_rank = {"summary": 3, "main": 2, "orphan": 1, "pec": 0, "anc": 0}

    def _store(code: str, name: str, credit: Optional[float], semester: str, source: str) -> bool:
        name = _clean_name(name)
        if not name or credit is None or credit <= 0:
            return False
        if name.lower() == "total" or PLACEHOLDER_NAME_RE.match(name):
            return False
        existing = subjects.get(code)
        if existing is not None and source_rank[existing["_source"]] >= source_rank[source]:
            return False
        subjects[code] = {
            "code": code,
            "name": name,
            "credits": float(credit),
            "semester": semester or (existing["semester"] if existing else ""),
            "department": infer_department_from_code(code),
            "_source": source,
        }
        return True

    def _next_meaningful(start: int, lookahead: int = 3):
        for j in range(start, min(n, start + lookahead)):
            l = all_lines[j]
            if not l:
                continue
            if SEM_HEADING_RE.match(l):
                return "heading", j
            if SUMMARY_LINE_RE.match(l) or MAIN_LINE_RE.match(l):
                return "row", j
            if ORPHAN_LINE_RE.match(l) or ORPHAN_SUMMARY_LINE_RE.match(l):
                return "orphan", j
            if FRAGMENT_BLACKLIST_RE.search(l) or re.search(CODE_PATTERN, l) or re.search(r'[A-Z]{2,6}XXX+\b', l):
                return "boundary", j
        return None, None

    current_sem = ""
    pending_fragments: List[str] = []
    last_stored_code: Optional[str] = None
    last_stored_line: Optional[int] = None
    in_pec_section = False
    pec_context_sem: Optional[str] = None
    pec_position = 0
    in_anc_section = False
    anc_position = 0

    i = 0
    while i < n:
        line = all_lines[i]
        if not line:
            i += 1
            continue

        if ANC_HEADING_RE.search(line):
            in_anc_section = True
            anc_position = 0
            in_pec_section = False
            pec_context_sem = None
            pending_fragments = []
            last_stored_code = None
            last_stored_line = None
            i += 1
            continue

        if PEC_SECTION_START_RE.match(line):
            in_pec_section = True
            pec_context_sem = None
            pec_position = 0
            pending_fragments = []
            last_stored_code = None
            last_stored_line = None
            i += 1
            continue

        if in_pec_section and line in ("V", "VI", "VII"):
            pec_context_sem = line
            i += 1
            continue

        if in_pec_section:
            line = re.sub(rf'^Professional\s+(?={CODE_PATTERN})', '', line)

        pec_header_m = PEC_HEADER_PREFIX_RE.match(line)
        if pec_header_m:
            label = pec_header_m.group("label").upper()
            pec_context_sem = PEC_LABEL_TO_SEM.get(label, pec_context_sem)
            in_pec_section = True
            line = line[pec_header_m.end():].lstrip('/ ').strip()
            pending_fragments = []
            if not line:
                i += 1
                continue

        heading_m = SEM_HEADING_RE.match(line)
        if heading_m:
            current_sem = heading_m.group(1)
            in_pec_section = False
            pec_context_sem = None
            in_anc_section = False
            anc_position = 0
            pending_fragments = []
            last_stored_code = None
            last_stored_line = None
            i += 1
            continue

        m = SUMMARY_LINE_RE.match(line)
        if m and m.group("sem") in VALID_ROMANS:
            credit = None
            try:
                credit = float(m.group("credit"))
            except ValueError:
                pass
            _store(m.group("code"), m.group("name"), credit, m.group("sem"), "summary")
            pending_fragments = []
            last_stored_code = None
            last_stored_line = None
            i += 1
            continue

        m = MAIN_LINE_RE.match(line)
        if m:
            code = m.group("code")
            name = m.group("name")
            credit = _last_number(m.group("rest"))
            wrote = _store(code, name, credit, current_sem, "main")
            pending_fragments = []
            kind, _ = _next_meaningful(i + 1)
            last_stored_code = code if wrote and kind not in ("orphan",) else None
            last_stored_line = i if last_stored_code else None
            i += 1
            continue

        m = ORPHAN_LINE_RE.match(line)
        if m:
            code = m.group("code")
            credit = _last_number(m.group("rest"))
            name = " ".join(pending_fragments).strip()
            pending_fragments = []
            last_stored_code = None
            last_stored_line = None
            if i + 1 < n and _is_plain_fragment(all_lines[i + 1]):
                kind, _ = _next_meaningful(i + 2)
                if kind != "orphan":
                    name = (name + " " + all_lines[i + 1]).strip()
                    i += 1
            _store(code, name, credit, current_sem, "orphan")
            i += 1
            continue

        m = ORPHAN_SUMMARY_LINE_RE.match(line)
        if m:
            code = m.group("code")
            credit = None
            try:
                credit = float(m.group("credit"))
            except (TypeError, ValueError):
                pass
            sem = m.group("sem") or current_sem
            name = " ".join(pending_fragments).strip()
            pending_fragments = []
            last_stored_code = None
            last_stored_line = None
            if i + 1 < n and _is_plain_fragment(all_lines[i + 1]):
                kind, _ = _next_meaningful(i + 2)
                if kind != "orphan":
                    name = (name + " " + all_lines[i + 1]).strip()
                    i += 1
            _store(code, name, credit, sem, "orphan")
            i += 1
            continue

        if in_pec_section:
            m = PEC_LINE_RE.match(line)
            if m:
                code = m.group("code")
                name = m.group("name")
                row_sem = m.group("sem")
                if row_sem:
                    pec_context_sem = row_sem
                pec_position += 1
                semester = row_sem or pec_context_sem or _pec_position_sem(pec_position)
                credit = 4.0
                wrote = _store(code, name, credit, semester, "pec")
                pending_fragments = []
                kind, _ = _next_meaningful(i + 1)
                last_stored_code = code if wrote and kind not in ("orphan",) else None
                last_stored_line = i if last_stored_code else None
                i += 1
                continue

        m = ANC_LINE_RE.match(line)
        used_prefixed_anc_match = False
        if not m and in_anc_section:
            m = ANC_LINE_PREFIXED_RE.match(line)
            used_prefixed_anc_match = m is not None
        if m and not re.search(r'[a-z]', m.group("name")):
            m = None
        if m:
            code = m.group("code")
            name = m.group("name")
            if in_anc_section:
                anc_position += 1
                semester = _anc_position_sem(anc_position)
            else:
                semester = ""
            credit = 3.0
            wrote = _store(code, name, credit, semester, "anc")
            pending_fragments = []
            kind, _ = _next_meaningful(i + 1)
            last_stored_code = code if wrote and kind not in ("orphan",) and not used_prefixed_anc_match else None
            last_stored_line = i if last_stored_code else None
            i += 1
            continue

        if FRAGMENT_BLACKLIST_RE.search(line):
            if re.search(r'Ancillary|Courses? [Oo]ffered|COURSES? OFFERED|Other Department Courses|Offered to\b', line):
                in_pec_section = False
                pec_context_sem = None
                in_anc_section = False
                anc_position = 0
            pending_fragments = []
            last_stored_code = None
            last_stored_line = None
            i += 1
            continue

        if _is_plain_fragment(line):
            kind, _ = _next_meaningful(i + 1)
            if kind == "orphan":
                pending_fragments = [line]
            elif last_stored_code is not None and last_stored_line == i - 1:
                merge_text = line
                if ')' in merge_text and '(' not in merge_text:
                    merge_text = merge_text.rsplit(')', 1)[-1].strip()
                existing = subjects.get(last_stored_code)
                if in_pec_section and existing is not None and existing.get("_source") == "pec":
                    sem_m = re.match(r'^(.*?)\s+(V|VI|VII)$', line)
                    if sem_m:
                        merge_text = sem_m.group(1).strip()
                        pec_context_sem = sem_m.group(2)
                        existing["semester"] = sem_m.group(2)
                if existing is not None and merge_text:
                    existing["name"] = _clean_name(existing["name"] + " " + merge_text)
                last_stored_code = None
                last_stored_line = None
            else:
                last_stored_code = None
                last_stored_line = None
                pending_fragments = [line]
        else:
            pending_fragments = []
            last_stored_code = None
            last_stored_line = None

        i += 1

    for sub in subjects.values():
        sub.pop("_source", None)
    return list(subjects.values())

# ==========================================
# STUDENT EXCEL PARSER
# ==========================================
_REPEATER_PATTERN = re.compile(r'[\s\-\(]*repeater[\)\s]*$', re.I)

def _extract_repeater(name: str) -> Tuple[str, bool]:
    is_repeater = bool(_REPEATER_PATTERN.search(name))
    clean = _REPEATER_PATTERN.sub('', name).strip()
    clean = re.sub(r'\s+', ' ', clean)
    return clean, is_repeater

def _clean_cell(value) -> str:
    if value is None:
        return ""
    try:
        if isinstance(value, float) and pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()

def _find_header_row(rows: List[List[Any]]) -> Optional[int]:
    for i, row in enumerate(rows[:15]):
        has_reg = any(re.search(r'Reg\.?\s*(?:No|Number)', _clean_cell(c), re.I) for c in row)
        has_name = any(re.search(r'(?:Name\s*of\s*(?:the)?\s*Students?|Students?\s*[_\s]*Name|Name\s*of\s*the\s*Candidate)', _clean_cell(c), re.I) for c in row)
        if has_reg and has_name:
            return i
    return None

def _department_from_sheet(sheet_name: str) -> str:
    sheet_name = re.sub(r'^\d{1,2}\.\d{1,2}\.\d{4}\s*', '', sheet_name)
    return re.sub(r'[-_]\s*[A-Z]$', '', sheet_name, flags=re.I).strip()

_DEPARTMENT_CANON: List[Tuple[str, Tuple[str, ...]]] = [
    ("CE",  (r'\bcivil\b', r'^ce$')),
    ("CHE", (r'\bchem', r'^che$', r'^chi$')),
    ("CSE", (r'computer\s*sci', r'^cse$')),
    ("DS",  (r'data\s*sci', r'^dat\s*sci$', r'^ds$')),
    ("ECE", (r'electronics.*communication', r'^ece$')),
    ("EEE", (r'electrical', r'^eee$')),
    ("EIE", (r'instrumentation', r'^eie$')),
    ("ENV", (r'environmental', r'^env\s*eng$', r'^env$')),
    ("CYS", (r'cyber\s*sec', r'information\s*sec', r'^info\s*sec$', r'^cys$')),
    ("IOT", (r'internet\s*of\s*things', r'^int\s*of\s*thi$', r'^iot$')),
    ("IT",  (r'information\s*tech', r'^it$')),
    ("MT",  (r'mechatronic', r'^mt$')),
    ("ME",  (r'mechanical', r'^me$')),
]

def _normalize_department(raw: str) -> str:
    d = (raw or "").strip().lower()
    if not d:
        return ""
    for code, patterns in _DEPARTMENT_CANON:
        for pat in patterns:
            if re.search(pat, d, re.I):
                return code
    return ""

def _section_from_sheet(sheet_name: str) -> str:
    sheet_name = re.sub(r'^\d{1,2}\.\d{1,2}\.\d{4}\s*', '', sheet_name)
    m = re.search(r'[-_]\s*([A-Z])$', sheet_name.strip(), flags=re.I)
    return m.group(1).upper() if m else ""

_SECTION_VALUE_PATTERN = re.compile(r'^[A-Z]$', re.I)

def _normalize_section(value: str) -> str:
    v = (value or "").strip().upper()
    return v if _SECTION_VALUE_PATTERN.match(v) else ""

def parse_students_excel(file_content: bytes) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
    students = []
    warnings: List[str] = []

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == "index":
            continue
        ws = wb[sheet_name]
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not all_rows:
            continue

        header_idx = _find_header_row(all_rows)
        if header_idx is None:
            continue
        header_row = all_rows[header_idx]

        reg_col = name_col = dept_col = programme_col = batch_col = section_col = None
        for col_idx, raw in enumerate(header_row):
            col_str = _clean_cell(raw)
            if not col_str:
                continue
            if re.search(r'Reg\.?\s*(?:No|Number)', col_str, re.I):
                reg_col = col_idx
            elif re.search(r'(?:Name\s*of\s*(?:the)?\s*Students?|Students?\s*[_\s]*Name|Name\s*of\s*the\s*Candidate)', col_str, re.I):
                name_col = col_idx
            elif re.search(r'Department|Dept', col_str, re.I):
                dept_col = col_idx
            elif re.search(r'Programme', col_str, re.I):
                programme_col = col_idx
            elif re.search(r'Section', col_str, re.I):
                section_col = col_idx
            elif re.search(r'Batch', col_str, re.I):
                batch_col = col_idx

        if reg_col is None or name_col is None:
            continue

        data_rows = all_rows[header_idx + 1:]

        def cell(row, idx):
            return _clean_cell(row[idx]) if idx is not None and idx < len(row) else ""

        n_cols = len(header_row)
        sample = [r for r in data_rows if any(_clean_cell(c) for c in r)][:30]
        batch_pattern = re.compile(r'^\d{4}\s*-\s*\d{4}$')
        if batch_col is None or section_col is None:
            for idx in range(n_cols):
                if idx in (reg_col, name_col, dept_col, programme_col):
                    continue
                vals = [cell(r, idx) for r in sample]
                vals = [v for v in vals if v]
                if not vals:
                    continue
                if batch_col is None and all(batch_pattern.match(v) for v in vals):
                    batch_col = idx
                elif section_col is None and all(_SECTION_VALUE_PATTERN.match(v) for v in vals):
                    section_col = idx

        if programme_col is None and dept_col is not None and dept_col + 1 < n_cols \
                and dept_col + 1 not in (batch_col, section_col):
            programme_col = dept_col + 1

        fallback_department_raw = _department_from_sheet(sheet_name)
        fallback_department = _normalize_department(fallback_department_raw) or fallback_department_raw
        fallback_section = _section_from_sheet(sheet_name)
        section_is_sheet_authoritative = bool(fallback_section)
        department_is_sheet_authoritative = bool(fallback_department) and bool(
            _normalize_department(fallback_department_raw)
        )

        last_seen_section = ""
        last_seen_department = ""
        sheet_reg_nos_seen = set()

        for row in data_rows:
            reg_no = cell(row, reg_col)
            name = cell(row, name_col)
            if not reg_no or not name or not re.search(r'\d', reg_no):
                continue

            name, is_repeater = _extract_repeater(name)
            if not name:
                continue

            row_department_raw = cell(row, dept_col)
            row_department_canon = _normalize_department(row_department_raw)

            if department_is_sheet_authoritative:
                department = fallback_department
            else:
                department = (
                    row_department_canon
                    or row_department_raw
                    or fallback_department
                )
            if department:
                last_seen_department = department
            elif last_seen_department:
                department = last_seen_department

            programme = cell(row, programme_col)
            if programme.strip().lower() in ("programme", "specialization"):
                programme = ""
            batch = cell(row, batch_col)

            if section_is_sheet_authoritative:
                section = fallback_section
            else:
                row_section = _normalize_section(cell(row, section_col))
                if row_section:
                    section = row_section
                    last_seen_section = row_section
                else:
                    section = last_seen_section

            sheet_reg_nos_seen.add(reg_no)

            students.append({
                "reg_no": reg_no,
                "name": name,
                "department": department,
                "programme": programme,
                "batch": batch,
                "section": section,
                "is_repeater": is_repeater
            })

    seen_global: Dict[str, Dict[str, Any]] = {}
    deduped: List[Dict[str, Any]] = []
    for s in students:
        prior = seen_global.get(s["reg_no"])
        if prior is None:
            seen_global[s["reg_no"]] = s
            deduped.append(s)
    students = deduped

    real_programmes = [s["programme"] for s in students if s["programme"]]
    if real_programmes:
        default_programme = Counter(real_programmes).most_common(1)[0][0]
        for s in students:
            if not s["programme"]:
                s["programme"] = default_programme

    parse_students_excel.last_warnings = warnings
    return students

# ==========================================
# RESULT PARSER (PDF)
# ==========================================
def parse_grade(grade_str: str) -> float:
    """Return the grade point for a raw grade string."""
    normalized = normalize_grade(grade_str or "")
    return GRADE_POINTS.get(normalized, 0.0)

_DEPT_HEADER_RE = re.compile(
    r'PTU\s*-\s*(?:'
    r'B\.?Tech-?BTech\s*-\s*([A-Za-z]+)'
    r'|([A-Za-z]+)\s*-\s*B\.?Tech\b'
    r'|B\.?Tech\s*-\s*([A-Za-z]+)\b'
    r'|(?:M\.?Tech|MCA)\s*-\s*([A-Za-z][A-Za-z .&]*?)(?=\s{2,}|\n|$)'
    r')',
    re.I
)

def match_department_header(text: str) -> Optional[str]:
    m = _DEPT_HEADER_RE.search(text)
    if not m:
        return None
    dept = m.group(1) or m.group(2) or m.group(3) or m.group(4)
    return dept.strip().upper()

_PROGRAMME_HEADER_RE = re.compile(
    r'PTU\s*-\s*(?:'
    r'(B\.?\s?Tech-?BTech|B\.?\s?Tech|M\.?\s?Tech|MCA)\b'
    r'|[A-Za-z]+\s*-\s*(B\.?\s?Tech|M\.?\s?Tech|MCA)\b'
    r')',
    re.I
)

def match_programme_header(text: str) -> Optional[str]:
    m = _PROGRAMME_HEADER_RE.search(text)
    if not m:
        return None
    prog = m.group(1) or m.group(2)
    prog = re.sub(r'[.\s]', '', prog).upper()
    if prog.endswith("BTECH") and prog != "BTECH":
        prog = "BTECH"
    return prog

def extract_semester_year_department(text: str) -> Dict[str, Optional[str]]:
    semester = None
    year = None
    sem_match = re.search(r'([IVXLCDM]+)\s*Semester', text, re.I)
    if sem_match:
        semester = sem_match.group(1).strip()
    year_match = re.search(r'Year\s*:\s*([A-Za-z]+\s*[-–]\s*\d{4})', text)
    if year_match:
        year = year_match.group(1).strip()
    department = match_department_header(text)
    programme = match_programme_header(text)
    return {"semester": semester, "year": year, "department": department, "programme": programme}

def extract_table_from_pdf(file_content: bytes) -> Dict[str, Any]:
    rows = []
    semester = None
    year = None
    department = None
    skipped_non_btech = 0
    current_department = None
    current_programme = None
    saw_any_text = False

    with pdfplumber.open(io.BytesIO(file_content)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            saw_any_text = True
            meta = extract_semester_year_department(text)
            if meta["semester"] and semester is None:
                semester = meta["semester"]
            if meta["year"] and year is None:
                year = meta["year"]
            if meta["department"]:
                current_department = meta["department"]
                if department is None and meta["programme"] == "BTECH":
                    department = current_department
            if meta["programme"]:
                current_programme = meta["programme"]

            for line in text.split('\n'):
                line = line.strip()
                if not line:
                    continue
                line_dept = match_department_header(line)
                if line_dept:
                    current_department = line_dept
                line_prog = match_programme_header(line)
                if line_prog:
                    current_programme = line_prog

                reg_match = re.search(r'\b(\d{10})\b', line)
                if not reg_match:
                    continue
                reg_no = reg_match.group(1)

                # Match subject codes and their grades.
                # Grade pattern includes lowercase letters to catch "Ab" (Absent) from PDFs.
                # Without the lowercase match, "Ab" would be truncated to "A" (a pass grade).
                pairs = re.findall(r'([A-Z]{2,4}\d{3,4})\s*-\s*([A-Za-z+]+)', line)
                if not pairs:
                    continue

                if current_programme and current_programme != "BTECH":
                    skipped_non_btech += 1
                    continue

                # Normalize every grade: Ab/AB/ABSENT → "F", passing grades uppercased
                grades = {code: normalize_grade(grade) for code, grade in pairs}
                first_subject = re.search(r'([A-Z]{2,4}\d{3,4})', line)
                if first_subject:
                    start_name = reg_match.end()
                    end_name = first_subject.start()
                    name = line[start_name:end_name].strip()
                else:
                    name = ""

                if not name:
                    parts = line.split()
                    try:
                        idx = parts.index(reg_no)
                        name_parts = []
                        for token in parts[idx+1:]:
                            if re.match(r'[A-Z]{2,4}\d{3,4}', token):
                                break
                            name_parts.append(token)
                        name = " ".join(name_parts)
                    except ValueError:
                        name = ""

                if name and grades:
                    rows.append({
                        "reg_no": reg_no,
                        "name": name,
                        "department": current_department or department,
                        "grades": grades
                    })

    return {
        "semester": semester,
        "year": year,
        "department": department,
        "rows": rows,
        "skipped_non_btech": skipped_non_btech,
    }

_REEVAL_REGNO_PATTERN = re.compile(r'^(\d{9,10}|\d{2}[A-Z]{2,4}\d{3,4})$')
_REEVAL_SUBCODE_PATTERN = re.compile(r'^[A-Z]{2,6}\d{2,4}$')
_REEVAL_GRADE_PATTERN = re.compile(r'^[A-Z]{1,3}\+?$')
_REEVAL_NON_BTECH_PROGRAMMES = {"MTECH", "MCA", "MSC", "PHD"}

def extract_reevaluation_table_from_pdf(file_content: bytes) -> Dict[str, Any]:
    rows_by_reg: Dict[str, Dict[str, Any]] = {}
    skipped_nc = 0
    skipped_non_btech = 0

    with pdfplumber.open(io.BytesIO(file_content)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.split('\n'):
                parts = line.strip().split()
                if "PTU" not in parts:
                    continue
                ptu_idx = parts.index("PTU")
                tail = parts[ptu_idx + 1:]
                if len(tail) < 4:
                    continue

                course = tail[0]
                fields = tail[1:]
                if len(fields) == 4:
                    _branch, reg_no, subcode, grade = fields
                elif len(fields) == 3:
                    _branch = None
                    reg_no, subcode, grade = fields
                else:
                    continue

                course_norm = course.upper().replace('.', '')
                if course_norm in _REEVAL_NON_BTECH_PROGRAMMES:
                    skipped_non_btech += 1
                    continue
                if course_norm != "BTECH":
                    continue

                if not _REEVAL_REGNO_PATTERN.match(reg_no) or not _REEVAL_SUBCODE_PATTERN.match(subcode) or not _REEVAL_GRADE_PATTERN.match(grade.upper()):
                    continue
                if grade.strip().upper() == "NC":
                    skipped_nc += 1
                    continue

                entry = rows_by_reg.setdefault(reg_no, {
                    "reg_no": reg_no,
                    "department": _branch,
                    "grades": {},
                })
                entry["grades"][subcode] = grade

    return {
        "rows": list(rows_by_reg.values()),
        "skipped_nc": skipped_nc,
        "skipped_non_btech": skipped_non_btech,
    }

# ==========================================
# UPLOAD ENDPOINTS
# ==========================================

# 1. Subjects Upload -> Dedicated subjects.db
@app.post("/upload/subjects", response_model=UploadResponse)
async def upload_subjects(
    file: UploadFile = File(...),
    db: Session = Depends(get_subjects_db),
    user: User = Depends(allow_write)
):
    content = await file.read()
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files accepted for subjects.")
    subjects_data = parse_subjects_pdf(content)
    if not subjects_data:
        raise HTTPException(status_code=400, detail="No subjects found in PDF.")
    errors = []
    added = 0
    updated = 0
    for sub in subjects_data:
        try:
            existing = db.query(Subject).filter(Subject.code == sub["code"]).first()
            if existing:
                existing.name = sub["name"]
                existing.credits = sub["credits"]
                existing.semester = sub["semester"] or existing.semester
                existing.department = sub["department"] or existing.department
                updated += 1
            else:
                new_sub = Subject(
                    code=sub["code"],
                    name=sub["name"],
                    credits=sub["credits"],
                    semester=sub["semester"] or "",
                    department=sub["department"]
                )
                db.add(new_sub)
                added += 1
            db.commit()
        except Exception as e:
            errors.append(f"Error processing subject {sub['code']}: {str(e)}")
            db.rollback()
    return UploadResponse(
        message=f"Subjects processed in Subjects DB. Added {added}, updated {updated}.",
        students_added=0,
        results_added=0,
        errors=errors
    )

# 2. Students Upload -> Dedicated per-batch database for each batch
@app.post("/upload/students", response_model=UploadResponse)
async def upload_students(
    file: UploadFile = File(...),
    user: User = Depends(allow_write)
):
    content = await file.read()
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files accepted for students.")
    students_data = parse_students_excel(content)
    if not students_data:
        raise HTTPException(status_code=400, detail="No students found in Excel.")

    # Group students by batch to write each to its dedicated batch DB
    by_batch: Dict[str, List[Dict[str, Any]]] = {}
    for stu in students_data:
        batch_key = stu.get("batch") or "General"
        by_batch.setdefault(batch_key, []).append(stu)

    errors = list(getattr(parse_students_excel, "last_warnings", []) or [])
    total_added = 0
    total_updated = 0

    for batch_name, batch_students in by_batch.items():
        batch_db = get_batch_session(batch_name)
        try:
            for stu in batch_students:
                try:
                    existing = batch_db.query(Student).filter(Student.reg_no == stu["reg_no"]).first()
                    if existing:
                        existing.name = stu["name"]
                        existing.department = stu["department"] or existing.department
                        existing.programme = stu["programme"] or existing.programme
                        existing.batch = stu["batch"] or existing.batch
                        existing.section = stu["section"] or existing.section
                        existing.is_repeater = stu.get("is_repeater", False)
                        existing.source = "roster"
                        existing.email = stu["reg_no"] + "@ptuniv.edu.in"
                        total_updated += 1
                    else:
                        new_stu = Student(
                            reg_no=stu["reg_no"],
                            name=stu["name"],
                            department=stu["department"] or "",
                            programme=stu["programme"] or "",
                            batch=stu["batch"] or batch_name,
                            section=stu["section"] or "",
                            is_repeater=stu.get("is_repeater", False),
                            source="roster",
                            email=stu["reg_no"] + "@ptuniv.edu.in"
                        )
                        batch_db.add(new_stu)
                        total_added += 1
                except Exception as e:
                    errors.append(f"Batch {batch_name}: error adding student {stu['reg_no']}: {str(e)}")
            batch_db.commit()
        finally:
            batch_db.close()

    return UploadResponse(
        message=f"Students processed into isolated batch database(s). Added {total_added}, updated {total_updated}.",
        students_added=total_added,
        results_added=0,
        errors=errors
    )

# 3. Results Upload -> Target batch database with subject validation from subjects.db
@app.post("/upload/results", response_model=UploadResponse)
async def upload_results(
    file: UploadFile = File(...),
    semester: Optional[str] = Form(None),
    batch: Optional[str] = Form(None),
    subjects_db: Session = Depends(get_subjects_db),
    user: User = Depends(allow_write)
):
    content = await file.read()
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files accepted for results.")

    parsed = extract_table_from_pdf(content)
    if semester:
        parsed['semester'] = semester

    errors = []
    results_added = 0

    if not parsed.get('semester'):
        errors.append("Semester could not be determined. Please provide it.")
    if not parsed.get('rows'):
        errors.append("No student data found in PDF.")
    if errors:
        raise HTTPException(status_code=400, detail="; ".join(errors))

    skipped_non_btech = parsed.get('skipped_non_btech', 0)
    if skipped_non_btech:
        errors.append(f"Skipped {skipped_non_btech} MTech/MCA result row(s); only BTech students are stored.")

    sem = parsed['semester']
    yr = parsed.get('year')
    skipped_unknown_students = 0

    # Build subjects lookup map from subjects.db
    all_subjects = subjects_db.query(Subject).all()
    subjects_by_code: Dict[str, List[Subject]] = {}
    for s in all_subjects:
        subjects_by_code.setdefault(s.code, []).append(s)

    # Cache open batch sessions during upload
    active_sessions: Dict[str, Session] = {}

    try:
        for row in parsed['rows']:
            reg_no = row['reg_no']
            grades = row['grades']

            # Determine target batch DB
            target_batch = batch
            if not target_batch:
                found = find_student_batch(reg_no)
                if found:
                    target_batch = found[0]

            if not target_batch:
                skipped_unknown_students += 1
                continue

            sanitized_key = sanitize_batch_name(target_batch)
            if sanitized_key not in active_sessions:
                active_sessions[sanitized_key] = get_batch_session(sanitized_key)
            batch_db = active_sessions[sanitized_key]

            student = batch_db.query(Student).filter(Student.reg_no == reg_no).first()
            if not student:
                # Check if student exists in another batch DB
                found = find_student_batch(reg_no)
                if found and found[1] != sanitized_key:
                    alt_key = found[1]
                    if alt_key not in active_sessions:
                        active_sessions[alt_key] = get_batch_session(alt_key)
                    batch_db = active_sessions[alt_key]
                    student = batch_db.query(Student).filter(Student.reg_no == reg_no).first()

            if not student:
                skipped_unknown_students += 1
                continue

            for code, grade_str in grades.items():
                candidates = subjects_by_code.get(code, [])
                if not candidates:
                    errors.append(f"Subject {code} not found in Subjects DB. Please add it first.")
                    continue
                elif len(candidates) == 1:
                    subject = candidates[0]
                else:
                    candidate_ids = [s.id for s in candidates]
                    prior = batch_db.query(Result).filter(
                        Result.student_id == student.id,
                        Result.subject_id.in_(candidate_ids),
                    ).first()
                    if prior:
                        subject = next(s for s in candidates if s.id == prior.subject_id)
                    else:
                        subject = candidates[0]

                gp = parse_grade(grade_str)
                stored_grade = normalize_grade(grade_str)       # AB/ABSENT → "F", others normalized
                is_fail_now = gp == 0.0                         # any zero-point grade is a fail
                result_sem = (sem or subject.semester or "I").strip()

                # Look up existing attempts for this student and subject
                existing_attempts = batch_db.query(Result).filter(
                    Result.student_id == student.id,
                    Result.subject_id == subject.id,
                ).order_by(Result.attempt.asc()).all()

                # Check if there is already an attempt recorded for this exact semester
                existing_same_sem = next((r for r in existing_attempts if r.semester == result_sem), None)

                if existing_same_sem:
                    # Update existing record for this semester
                    existing_same_sem.grade = stored_grade
                    existing_same_sem.grade_point = gp
                    existing_same_sem.year = yr or existing_same_sem.year
                    if target_batch:
                        existing_same_sem.batch = target_batch
                    existing_same_sem.had_arrear = bool(existing_same_sem.had_arrear) or is_fail_now or (len(existing_attempts) > 1)
                else:
                    # New attempt in a different semester
                    max_attempt = max([r.attempt for r in existing_attempts], default=0)
                    new_attempt = max_attempt + 1
                    had_arr = is_fail_now or (new_attempt > 1) or any(r.had_arrear or r.grade_point == 0 for r in existing_attempts)
                    result = Result(
                        student_id=student.id,
                        subject_id=subject.id,
                        semester=result_sem,
                        year=yr,
                        grade=stored_grade,
                        grade_point=gp,
                        batch=target_batch or student.batch,
                        attempt=new_attempt,
                        had_arrear=had_arr
                    )
                    batch_db.add(result)
                    results_added += 1

        # Commit all modified batch DB sessions
        for s in active_sessions.values():
            s.commit()
    finally:
        for s in active_sessions.values():
            s.close()

    if skipped_unknown_students:
        errors.append(
            f"Skipped {skipped_unknown_students} student record(s) not found in any batch database."
        )

    return UploadResponse(
        message=f"Results processed in batch database(s). Results saved: {results_added}.",
        students_added=0,
        results_added=results_added,
        errors=errors
    )

# 4. Re-evaluation Upload -> Target batch database
@app.post("/upload/reevaluation", response_model=UploadResponse)
async def upload_reevaluation(
    file: UploadFile = File(...),
    semester: Optional[str] = Form(None),
    batch: Optional[str] = Form(None),
    subjects_db: Session = Depends(get_subjects_db),
    user: User = Depends(allow_write)
):
    content = await file.read()
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only PDF files accepted for re-evaluation results.")

    parsed = extract_reevaluation_table_from_pdf(content)
    errors = []
    results_updated = 0

    if not parsed.get('rows'):
        errors.append("No re-evaluation rows found in PDF.")
    if errors:
        raise HTTPException(status_code=400, detail="; ".join(errors))

    all_subjects = subjects_db.query(Subject).all()
    subjects_by_code: Dict[str, List[Subject]] = {}
    for s in all_subjects:
        subjects_by_code.setdefault(s.code, []).append(s)

    active_sessions: Dict[str, Session] = {}

    try:
        for row in parsed['rows']:
            reg_no = row['reg_no']
            grades = row['grades']

            target_batch = batch
            if not target_batch:
                found = find_student_batch(reg_no)
                if found:
                    target_batch = found[0]

            if not target_batch:
                continue

            sanitized_key = sanitize_batch_name(target_batch)
            if sanitized_key not in active_sessions:
                active_sessions[sanitized_key] = get_batch_session(sanitized_key)
            batch_db = active_sessions[sanitized_key]

            student = batch_db.query(Student).filter(Student.reg_no == reg_no).first()
            if not student:
                continue

            for code, grade_str in grades.items():
                candidates = subjects_by_code.get(code, [])
                if not candidates:
                    continue
                subject = candidates[0]

                existing_query = batch_db.query(Result).filter(
                    Result.student_id == student.id,
                    Result.subject_id == subject.id,
                )
                if semester:
                    existing = existing_query.filter(Result.semester == semester).first()
                else:
                    existing = existing_query.order_by(Result.attempt.desc()).first()

                if not existing:
                    continue

                new_gp = parse_grade(grade_str)
                if new_gp <= existing.grade_point and existing.grade_point > 0:
                    continue

                existing.grade = normalize_grade(grade_str)
                existing.grade_point = new_gp
                results_updated += 1

        for s in active_sessions.values():
            s.commit()
    finally:
        for s in active_sessions.values():
            s.close()

    return UploadResponse(
        message="Re-evaluation processed",
        students_added=0,
        results_added=results_updated,
        errors=errors
    )

# ==========================================
# CRUD ROUTES: STUDENTS
# ==========================================
@app.get("/students", response_model=List[StudentResponse])
def get_students(
    department: Optional[str] = None,
    batch: Optional[str] = None,
    user: User = Depends(allow_all)
):
    if batch and batch.strip().lower() != "all":
        batch_db = get_batch_session(batch)
        try:
            query = batch_db.query(Student).filter(Student.source == "roster")
            if department:
                query = query.filter(Student.department == department)
            return query.all()
        finally:
            batch_db.close()

    # Query across all batch databases
    students: List[Student] = []
    for _key, batch_db in get_all_batch_sessions():
        try:
            query = batch_db.query(Student).filter(Student.source == "roster")
            if department:
                query = query.filter(Student.department == department)
            students.extend(query.all())
        finally:
            batch_db.close()
    return students

_VALID_BATCH_PATTERN = re.compile(r'^\d{4}\s*-\s*\d{4}$')

@app.get("/students/batches", response_model=List[str])
def get_student_batches(user: User = Depends(allow_all)):
    return get_all_batch_names()

@app.get("/batches/{batch_name}/info")
def get_batch_info(batch_name: str, user: User = Depends(allow_all)):
    cleaned = (batch_name or "").strip()
    if not cleaned:
        raise HTTPException(status_code=400, detail="Batch name is required")
    sanitized = sanitize_batch_name(cleaned)
    db_path = os.path.join(BATCHES_DIR, f"batch_{sanitized}.db")
    if not os.path.exists(db_path):
        alt_found = False
        if os.path.exists(BATCHES_DIR):
            for fname in os.listdir(BATCHES_DIR):
                if fname.startswith("batch_") and fname.endswith(".db"):
                    raw_key = fname[len("batch_"):-len(".db")]
                    if raw_key.lower() == sanitized.lower() or raw_key.replace('_', '-').lower() == cleaned.lower():
                        db_path = os.path.join(BATCHES_DIR, fname)
                        alt_found = True
                        break
        if not alt_found and not os.path.exists(db_path):
            raise HTTPException(status_code=404, detail=f"Batch '{batch_name}' not found")

    students_count = 0
    results_count = 0
    try:
        eng = create_engine(f"sqlite:///{db_path}", connect_args={"check_same_thread": False})
        with sessionmaker(bind=eng)() as s:
            students_count = s.query(Student).count()
            results_count = s.query(Result).count()
    except Exception as e:
        logger.warning("Error reading batch info: %s", e)

    return {
        "batch": cleaned,
        "students": students_count,
        "results": results_count
    }

@app.delete("/batches/{batch_name}")
def delete_batch(batch_name: str, user: User = Depends(allow_admin)):
    cleaned = (batch_name or "").strip()
    if not cleaned:
        raise HTTPException(status_code=400, detail="Batch name is required")
    
    deleted = delete_batch_database(cleaned)
    if not deleted:
        raise HTTPException(status_code=404, detail=f"Batch '{batch_name}' not found")
        
    return {
        "message": f"Batch '{batch_name}' and all associated students and results have been completely deleted."
    }

@app.post("/students", response_model=StudentResponse)
def create_student(student: StudentCreate, user: User = Depends(allow_write)):
    batch_db = get_batch_session(student.batch)
    try:
        existing = batch_db.query(Student).filter(Student.reg_no == student.reg_no).first()
        if existing:
            raise HTTPException(status_code=400, detail="Student with this reg_no already exists in this batch")
        db_student = Student(**student.model_dump())
        db_student.email = db_student.reg_no + "@ptuniv.edu.in"
        batch_db.add(db_student)
        batch_db.commit()
        batch_db.refresh(db_student)
        return db_student
    finally:
        batch_db.close()

@app.put("/students/by-reg/{reg_no}", response_model=StudentResponse)
def update_student_by_reg(reg_no: str, student: StudentCreate, user: User = Depends(allow_write)):
    cleaned_reg = (reg_no or "").strip().lower()
    found_info = find_student_batch(cleaned_reg)
    if not found_info:
        raise HTTPException(status_code=404, detail=f"Student with Reg No '{reg_no}' not found")

    current_batch_name, current_sanitized_key = found_info
    target_batch = (student.batch or current_batch_name).strip()

    # If student is being moved to another batch DB
    if sanitize_batch_name(target_batch) != current_sanitized_key:
        old_db = get_batch_session(current_sanitized_key)
        new_db = get_batch_session(target_batch)
        try:
            old_stu = old_db.query(Student).filter(func.lower(Student.reg_no) == cleaned_reg).first()
            if not old_stu:
                raise HTTPException(status_code=404, detail="Student not found in source batch")
            
            # Check if reg_no exists in target batch
            existing_target = new_db.query(Student).filter(func.lower(Student.reg_no) == cleaned_reg).first()
            if existing_target:
                raise HTTPException(status_code=400, detail="Student already exists in target batch")

            # Create in new batch DB
            new_stu = Student(
                reg_no=student.reg_no,
                name=student.name,
                department=student.department,
                programme=student.programme,
                batch=target_batch,
                section=student.section,
                is_repeater=student.is_repeater,
                source=old_stu.source,
                email=student.reg_no + "@ptuniv.edu.in"
            )
            new_db.add(new_stu)
            new_db.flush()

            # Move results
            old_results = old_db.query(Result).filter(Result.student_id == old_stu.id).all()
            for r in old_results:
                new_res = Result(
                    student_id=new_stu.id,
                    subject_id=r.subject_id,
                    semester=r.semester,
                    year=r.year,
                    grade=r.grade,
                    grade_point=r.grade_point,
                    batch=target_batch,
                    had_arrear=r.had_arrear
                )
                new_db.add(new_res)

            old_db.delete(old_stu)
            old_db.commit()
            new_db.commit()
            new_db.refresh(new_stu)
            return new_stu
        finally:
            old_db.close()
            new_db.close()

    # Same batch DB update
    batch_db = get_batch_session(current_sanitized_key)
    try:
        db_student = batch_db.query(Student).filter(func.lower(Student.reg_no) == cleaned_reg).first()
        if not db_student:
            raise HTTPException(status_code=404, detail="Student not found")
        for key, value in student.model_dump().items():
            setattr(db_student, key, value)
        db_student.email = db_student.reg_no + "@ptuniv.edu.in"
        batch_db.commit()
        batch_db.refresh(db_student)
        return db_student
    finally:
        batch_db.close()

@app.delete("/students/by-reg/{reg_no}")
def delete_student_by_reg(reg_no: str, user: User = Depends(allow_admin)):
    cleaned_reg = (reg_no or "").strip().lower()
    found_info = find_student_batch(cleaned_reg)
    if not found_info:
        raise HTTPException(status_code=404, detail=f"Student with Reg No '{reg_no}' not found")

    batch_db = get_batch_session(found_info[1])
    try:
        db_student = batch_db.query(Student).filter(func.lower(Student.reg_no) == cleaned_reg).first()
        if not db_student:
            raise HTTPException(status_code=404, detail="Student not found")
        batch_db.delete(db_student)
        batch_db.commit()
        return {"detail": "Student deleted successfully"}
    finally:
        batch_db.close()

@app.put("/students/{student_id}", response_model=StudentResponse)
def update_student(student_id: int, student: StudentCreate, batch: Optional[str] = None, user: User = Depends(allow_write)):
    # If reg_no is known, delegate to exact reg_no updater
    found_info = find_student_batch(student.reg_no)
    if found_info:
        return update_student_by_reg(student.reg_no, student, user=user)

    target_batch = student.batch or batch
    if target_batch:
        batch_db = get_batch_session(target_batch)
        try:
            db_student = batch_db.query(Student).filter(Student.id == student_id).first()
            if db_student:
                for key, value in student.model_dump().items():
                    setattr(db_student, key, value)
                db_student.email = db_student.reg_no + "@ptuniv.edu.in"
                batch_db.commit()
                batch_db.refresh(db_student)
                return db_student
        finally:
            batch_db.close()

    for _key, batch_db in get_all_batch_sessions():
        try:
            db_student = batch_db.query(Student).filter(Student.id == student_id).first()
            if db_student:
                for key, value in student.model_dump().items():
                    setattr(db_student, key, value)
                db_student.email = db_student.reg_no + "@ptuniv.edu.in"
                batch_db.commit()
                batch_db.refresh(db_student)
                return db_student
        finally:
            batch_db.close()

    raise HTTPException(status_code=404, detail="Student not found")

@app.delete("/students/{student_id}")
def delete_student(student_id: int, batch: Optional[str] = None, reg_no: Optional[str] = None, user: User = Depends(allow_admin)):
    if reg_no:
        return delete_student_by_reg(reg_no, user=user)

    if batch:
        batch_db = get_batch_session(batch)
        try:
            db_student = batch_db.query(Student).filter(Student.id == student_id).first()
            if db_student:
                batch_db.delete(db_student)
                batch_db.commit()
                return {"detail": "Student deleted"}
        finally:
            batch_db.close()

    for _key, batch_db in get_all_batch_sessions():
        try:
            db_student = batch_db.query(Student).filter(Student.id == student_id).first()
            if db_student:
                batch_db.delete(db_student)
                batch_db.commit()
                return {"detail": "Student deleted"}
        finally:
            batch_db.close()

    raise HTTPException(status_code=404, detail="Student not found")

# ==========================================
# CRUD ROUTES: SUBJECTS (Dedicated subjects.db)
# ==========================================
@app.get("/subjects", response_model=List[SubjectResponse])
def get_subjects(
    department: Optional[str] = None,
    semester: Optional[str] = None,
    db: Session = Depends(get_subjects_db),
    user: User = Depends(allow_all)
):
    query = db.query(Subject)
    if department:
        query = query.filter(Subject.department == department)
    if semester:
        query = query.filter(Subject.semester == semester)
    return query.all()

@app.post("/subjects", response_model=SubjectResponse)
def create_subject(subject: SubjectCreate, db: Session = Depends(get_subjects_db), user: User = Depends(allow_write)):
    existing = db.query(Subject).filter(Subject.code == subject.code).first()
    if existing:
        raise HTTPException(status_code=400, detail="Subject code already exists")
    db_subject = Subject(**subject.model_dump())
    db.add(db_subject)
    db.commit()
    db.refresh(db_subject)
    return db_subject

@app.put("/subjects/{subject_id}", response_model=SubjectResponse)
def update_subject(subject_id: int, subject: SubjectCreate, db: Session = Depends(get_subjects_db), user: User = Depends(allow_write)):
    db_subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if not db_subject:
        raise HTTPException(status_code=404, detail="Subject not found")
    for key, value in subject.model_dump().items():
        setattr(db_subject, key, value)
    db.commit()
    db.refresh(db_subject)
    return db_subject

@app.delete("/subjects/{subject_id}")
def delete_subject(subject_id: int, db: Session = Depends(get_subjects_db), user: User = Depends(allow_admin)):
    db_subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if not db_subject:
        raise HTTPException(status_code=404, detail="Subject not found")
    db.delete(db_subject)
    db.commit()
    return {"detail": "Subject deleted"}

# ==========================================
# CRUD ROUTES: RESULTS (Direct batch DB query + fast subjects map)
# ==========================================
@app.get("/results", response_model=List[ResultWithDetails])
def get_results(
    student_id: Optional[int] = None,
    reg_no: Optional[str] = None,
    department: Optional[str] = None,
    semester: Optional[str] = None,
    subject_code: Optional[str] = None,
    batch: Optional[str] = None,
    subjects_db: Session = Depends(get_subjects_db),
    user: User = Depends(allow_all)
):
    subjects_map = get_subjects_map(subjects_db)
    target_subject_ids = None
    if subject_code and subject_code.strip():
        sc = subject_code.strip().upper()
        exact_ids = [sid for sid, sub in subjects_map.items() if sub.code and sub.code.strip().upper() == sc]
        if exact_ids:
            target_subject_ids = set(exact_ids)
        else:
            partial_ids = [sid for sid, sub in subjects_map.items() if sub.code and sc in sub.code.strip().upper()]
            target_subject_ids = set(partial_ids)

        if not target_subject_ids:
            return []

    response: List[ResultWithDetails] = []

    # If batch is specified (or deduced from reg_no), query only that batch DB
    if batch and batch.strip().lower() != "all":
        batch_db = get_batch_session(batch)
        try:
            query = batch_db.query(Result).join(Student)
            if student_id:
                query = query.filter(Result.student_id == student_id)
            if reg_no:
                query = query.filter(Student.reg_no == reg_no)
            if department:
                query = query.filter(Student.department == department)
            if semester:
                query = query.filter(Result.semester == semester)
            if target_subject_ids is not None:
                query = query.filter(Result.subject_id.in_(target_subject_ids))

            for r in query.all():
                sub = subjects_map.get(r.subject_id)
                response.append(ResultWithDetails(
                    id=r.id,
                    student_id=r.student_id,
                    subject_id=r.subject_id,
                    semester=r.semester,
                    year=r.year,
                    grade=r.grade,
                    grade_point=r.grade_point,
                    batch=r.batch or batch,
                    attempt=getattr(r, 'attempt', 1) or 1,
                    had_arrear=bool(r.had_arrear),
                    student_name=r.student.name,
                    reg_no=r.student.reg_no,
                    department=r.student.department or "",
                    subject_code=sub.code if sub else f"SUB{r.subject_id}",
                    subject_name=sub.name if sub else "Unknown Subject",
                    credits=sub.credits if sub else 0.0
                ))
            return response
        finally:
            batch_db.close()

    # Query across all batch databases
    for batch_name, batch_db in get_all_batch_sessions():
        try:
            query = batch_db.query(Result).join(Student)
            if student_id:
                query = query.filter(Result.student_id == student_id)
            if reg_no:
                query = query.filter(Student.reg_no == reg_no)
            if department:
                query = query.filter(Student.department == department)
            if semester:
                query = query.filter(Result.semester == semester)
            if target_subject_ids is not None:
                query = query.filter(Result.subject_id.in_(target_subject_ids))

            for r in query.all():
                sub = subjects_map.get(r.subject_id)
                response.append(ResultWithDetails(
                    id=r.id,
                    student_id=r.student_id,
                    subject_id=r.subject_id,
                    semester=r.semester,
                    year=r.year,
                    grade=r.grade,
                    grade_point=r.grade_point,
                    batch=r.batch or batch_name,
                    attempt=getattr(r, 'attempt', 1) or 1,
                    had_arrear=bool(r.had_arrear),
                    student_name=r.student.name,
                    reg_no=r.student.reg_no,
                    department=r.student.department or "",
                    subject_code=sub.code if sub else f"SUB{r.subject_id}",
                    subject_name=sub.name if sub else "Unknown Subject",
                    credits=sub.credits if sub else 0.0
                ))
        finally:
            batch_db.close()

    return response

@app.post("/results", response_model=ResultResponse)
def create_result(
    result: ResultCreate,
    batch: Optional[str] = None,
    subjects_db: Session = Depends(get_subjects_db),
    user: User = Depends(allow_write)
):
    subject = subjects_db.query(Subject).filter(Subject.id == result.subject_id).first()
    if not subject:
        raise HTTPException(status_code=404, detail="Subject not found")

    target_batch = result.batch or batch
    if not target_batch:
        raise HTTPException(status_code=400, detail="Batch is required to add result")

    batch_db = get_batch_session(target_batch)
    try:
        student = batch_db.query(Student).filter(Student.id == result.student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="Student not found in this batch database")

        existing_attempts = batch_db.query(Result).filter(
            Result.student_id == result.student_id,
            Result.subject_id == result.subject_id,
        ).all()

        attempt_no = result.attempt or (len(existing_attempts) + 1)
        stored_grade = normalize_grade(result.grade or "")
        gp = parse_grade(result.grade or "")
        is_fail_now = gp == 0.0
        had_arr = is_fail_now or (attempt_no > 1) or any(r.had_arrear or r.grade_point == 0 for r in existing_attempts)

        db_result = Result(
            student_id=result.student_id,
            subject_id=result.subject_id,
            semester=result.semester,
            year=result.year,
            grade=stored_grade,
            grade_point=gp,
            batch=target_batch,
            attempt=attempt_no,
            had_arrear=had_arr
        )
        batch_db.add(db_result)
        batch_db.commit()
        batch_db.refresh(db_result)
        return db_result
    finally:
        batch_db.close()

@app.put("/results/{result_id}", response_model=ResultResponse)
def update_result(result_id: int, result: ResultUpdate, batch: Optional[str] = None, user: User = Depends(allow_write)):
    stored_grade = normalize_grade(result.grade or "")
    new_gp = parse_grade(result.grade or "")

    if batch:
        batch_db = get_batch_session(batch)
        try:
            db_result = batch_db.query(Result).filter(Result.id == result_id).first()
            if db_result:
                db_result.grade = stored_grade
                db_result.grade_point = new_gp
                if result.year is not None:
                    db_result.year = result.year
                if result.batch is not None:
                    db_result.batch = result.batch
                if result.attempt is not None:
                    db_result.attempt = result.attempt
                db_result.had_arrear = (new_gp == 0.0) or (db_result.attempt > 1)
                batch_db.commit()
                batch_db.refresh(db_result)
                return db_result
        finally:
            batch_db.close()

    for _key, batch_db in get_all_batch_sessions():
        try:
            db_result = batch_db.query(Result).filter(Result.id == result_id).first()
            if db_result:
                db_result.grade = stored_grade
                db_result.grade_point = new_gp
                if result.year is not None:
                    db_result.year = result.year
                if result.batch is not None:
                    db_result.batch = result.batch
                if result.attempt is not None:
                    db_result.attempt = result.attempt
                db_result.had_arrear = (new_gp == 0.0) or (db_result.attempt > 1)
                batch_db.commit()
                batch_db.refresh(db_result)
                return db_result
        finally:
            batch_db.close()

    raise HTTPException(status_code=404, detail="Result not found")

@app.delete("/results/{result_id}")
def delete_result(result_id: int, batch: Optional[str] = None, user: User = Depends(allow_admin)):
    if batch:
        batch_db = get_batch_session(batch)
        try:
            db_result = batch_db.query(Result).filter(Result.id == result_id).first()
            if db_result:
                batch_db.delete(db_result)
                batch_db.commit()
                return {"detail": "Result deleted"}
        finally:
            batch_db.close()

    for _key, batch_db in get_all_batch_sessions():
        try:
            db_result = batch_db.query(Result).filter(Result.id == result_id).first()
            if db_result:
                batch_db.delete(db_result)
                batch_db.commit()
                return {"detail": "Result deleted"}
        finally:
            batch_db.close()

    raise HTTPException(status_code=404, detail="Result not found")

# ==========================================
# SGPA / CGPA CALCULATION
# ==========================================
def _arrear_bucket(count: int) -> str:
    if count <= 0:
        return "0"
    if count == 1:
        return "1"
    if count == 2:
        return "2"
    return "3+"

@app.get("/grades/summary", response_model=List[GradeSummary])
def get_grade_summary(
    student_id: Optional[int] = None,
    reg_no: Optional[str] = None,
    department: Optional[str] = None,
    semester: Optional[str] = None,
    batch: Optional[str] = None,
    arrears: Optional[List[str]] = Query(None),
    subjects_db: Session = Depends(get_subjects_db),
    user: User = Depends(allow_all)
):
    subjects_map = get_subjects_map(subjects_db)
    credits_map = {sid: sub.credits for sid, sub in subjects_map.items()}

    # Collect batches to process
    batch_sessions = []
    if batch and batch.strip().lower() != "all":
        batch_sessions = [(batch, get_batch_session(batch))]
    else:
        batch_sessions = get_all_batch_sessions()

    all_summaries: List[GradeSummary] = []
    sem_label = semester if semester else "All"

    for _batch_name, batch_db in batch_sessions:
        try:
            stu_query = batch_db.query(
                Student.id, Student.reg_no, Student.name, Student.department
            ).filter(Student.source == "roster")
            if student_id:
                stu_query = stu_query.filter(Student.id == student_id)
            if reg_no:
                stu_query = stu_query.filter(Student.reg_no == reg_no)
            if department:
                stu_query = stu_query.filter(Student.department == department)

            students = stu_query.all()
            if not students:
                continue

            stu_ids = [s[0] for s in students]

            # Fetch results
            res_query = batch_db.query(
                Result.student_id, Result.subject_id, Result.grade_point, Result.had_arrear
            ).filter(Result.student_id.in_(stu_ids))
            if semester:
                res_query = res_query.filter(Result.semester == semester)
            res_rows = res_query.all()

            # Group results by student
            student_results: Dict[int, List[Tuple[int, float, bool]]] = {}
            for sid, sub_id, gp, had_arr in res_rows:
                student_results.setdefault(sid, []).append((sub_id, gp, had_arr))

            # Arrear counts (across all semesters for each student in this batch DB)
            arr_rows = (
                batch_db.query(Result.student_id, func.count(func.distinct(Result.subject_id)))
                .filter(Result.student_id.in_(stu_ids), Result.had_arrear.is_(True))
                .group_by(Result.student_id)
                .all()
            )
            arr_map = {s_id: cnt for s_id, cnt in arr_rows}

            for stu_id, stu_reg, stu_name, stu_dept in students:
                results_list = student_results.get(stu_id, [])
                total_credits = 0.0
                grade_points_sum = 0.0
                earned_credits = 0.0

                for sub_id, gp, _had_arr in results_list:
                    cr = credits_map.get(sub_id, 0.0)
                    total_credits += cr
                    grade_points_sum += (gp * cr)
                    if gp > 0:
                        earned_credits += cr

                gpa = round(grade_points_sum / total_credits, 2) if total_credits > 0 else None
                arrear_count = arr_map.get(stu_id, 0)

                # Filter by arrears
                if arrears and isinstance(arrears, (list, tuple, set)):
                    wanted_buckets = set()
                    skip_filter = False
                    for raw in arrears:
                        token = (raw or "").strip().lower()
                        if token in ("all", ""):
                            skip_filter = True
                            break
                        if token in ("0", "no", "none", "no arrears", "no_arrears"):
                            wanted_buckets.add("0")
                        elif token in ("1", "1 arrear", "1_arrear"):
                            wanted_buckets.add("1")
                        elif token in ("2", "2 arrear", "2_arrear", "2 arrears", "2_arrears"):
                            wanted_buckets.add("2")
                        else:
                            wanted_buckets.add("3+")
                    if not skip_filter and _arrear_bucket(arrear_count) not in wanted_buckets:
                        continue

                all_summaries.append(GradeSummary(
                    student_id=stu_id,
                    reg_no=stu_reg,
                    name=stu_name,
                    department=stu_dept,
                    semester=sem_label,
                    sgpa=gpa if semester else None,
                    cgpa=gpa if not semester else None,
                    total_credits=total_credits,
                    earned_credits=earned_credits,
                    grade_points_sum=grade_points_sum,
                    arrear_count=arrear_count
                ))
        finally:
            batch_db.close()

    return all_summaries

# ==========================================
# PUBLIC STATS (Homepage Ticker)
# ==========================================
@app.get("/stats/public")
def get_public_stats(subjects_db: Session = Depends(get_subjects_db)):
    subjects_count = subjects_db.query(Subject).count()
    students_count = 0
    results_count = 0
    depts_set = set()

    for _key, batch_db in get_all_batch_sessions():
        try:
            students_count += batch_db.query(Student).filter(Student.source == "roster").count()
            results_count += batch_db.query(Result).count()
            for row in batch_db.query(Student.department).distinct().all():
                if row[0]:
                    depts_set.add(row[0])
        finally:
            batch_db.close()

    depts_count = len(depts_set)
    return {
        "students": students_count,
        "subjects": subjects_count,
        "results": results_count,
        "departments": depts_count
    }

# ==========================================
# REPORT CARD (Public & Testing Direct Generator)
# ==========================================
_ROMAN_SEM_ORDER = {r: i for i, r in enumerate(
    ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]
)}

def _semester_sort_key(sem: str):
    return (_ROMAN_SEM_ORDER.get((sem or "").strip().upper(), 99), sem or "")

def _build_student_report_card(student: Student, batch_db: Session, subjects_db: Session) -> ReportCardResponse:
    results = batch_db.query(Result).filter(Result.student_id == student.id).order_by(Result.attempt.asc(), Result.id.asc()).all()
    if not results:
        raise HTTPException(status_code=404, detail="No results have been published for this student yet")

    subjects_map = get_subjects_map(subjects_db)

    # 1. Group results by subject_id to analyze attempt histories and arrear lifecycles
    subject_results: Dict[int, List[Result]] = {}
    for r in results:
        subject_results.setdefault(r.subject_id, []).append(r)

    arrear_history_list: List[ArrearSubjectHistory] = []
    has_any_arrears = False

    # Pre-analyze each subject for attempt/clearance tracking
    subject_meta: Dict[int, Dict[str, Any]] = {}
    for sub_id, att_list in subject_results.items():
        sub = subjects_map.get(sub_id)
        sub_code = sub.code if sub else f"SUB{sub_id}"
        sub_name = sub.name if sub else "Unknown Subject"
        sub_cr = sub.credits if sub else 0.0

        # Sort attempts chronologically
        sorted_atts = sorted(att_list, key=lambda x: (x.attempt or 1, _semester_sort_key(x.semester), x.id or 0))
        att_seq_map = {a.id: idx for idx, a in enumerate(sorted_atts, start=1)}

        # Identify failed attempts vs passed attempts
        failed_atts = [
            a for a in sorted_atts
            if a.grade_point == 0 or (a.grade or "").strip().upper() in ("F", "AB", "ABSENT", "NC", "E", "Z")
        ]
        passed_atts = [
            a for a in sorted_atts
            if a.grade_point > 0 and (a.grade or "").strip().upper() not in ("F", "AB", "ABSENT", "NC", "E", "Z")
        ]

        had_fail = len(failed_atts) > 0 or any(a.had_arrear for a in sorted_atts) or len(sorted_atts) > 1
        is_cleared = len(passed_atts) > 0
        cleared_att = passed_atts[-1] if is_cleared else None
        first_fail = failed_atts[0] if failed_atts else (sorted_atts[0] if had_fail else None)

        failed_sem = first_fail.semester if first_fail else (sorted_atts[0].semester if had_fail else None)
        failed_grd = first_fail.grade if first_fail else ("F" if had_fail else None)
        cleared_sem = cleared_att.semester if cleared_att else None
        cleared_grd = cleared_att.grade if cleared_att else None
        cleared_att_num = att_seq_map.get(cleared_att.id, 2) if cleared_att else None
        total_atts_count = len(sorted_atts)

        if had_fail:
            has_any_arrears = True
            status_text = (
                f"Cleared in Sem {cleared_sem} (Attempt {cleared_att_num} · Grade {cleared_grd})"
                if is_cleared else
                f"Arrear Pending (Attempt {total_atts_count} in Sem {failed_sem})"
            )
            arrear_history_list.append(ArrearSubjectHistory(
                subject_code=sub_code,
                subject_name=sub_name,
                credits=sub_cr,
                failed_semester=failed_sem or sorted_atts[0].semester,
                failed_grade=failed_grd or "F",
                cleared_semester=cleared_sem,
                cleared_grade=cleared_grd,
                total_attempts=total_atts_count,
                is_cleared=is_cleared,
                status=status_text
            ))

        # Store metadata for quick lookup when constructing semester tables
        hist_details = [
            AttemptDetail(
                attempt=att_seq_map.get(a.id, idx),
                semester=a.semester,
                year=a.year,
                grade=a.grade,
                grade_point=a.grade_point,
                is_cleared=(a.grade_point > 0 and (a.grade or "").strip().upper() not in ("F", "AB", "ABSENT", "NC", "E", "Z"))
            )
            for idx, a in enumerate(sorted_atts, start=1)
        ]

        subject_meta[sub_id] = {
            "had_fail": had_fail,
            "is_cleared": is_cleared,
            "failed_sem": failed_sem,
            "failed_grd": failed_grd,
            "cleared_sem": cleared_sem,
            "cleared_grd": cleared_grd,
            "cleared_attempt": cleared_att_num,
            "total_attempts": total_atts_count,
            "original_sem": sorted_atts[0].semester,
            "history": hist_details,
            "seq_map": att_seq_map
        }

    # 2. Group by semester for semester-wise breakdown
    by_sem: Dict[str, List[Result]] = {}
    for r in results:
        by_sem.setdefault(r.semester, []).append(r)

    semesters: List[ReportCardSemester] = []

    for sem in sorted(by_sem.keys(), key=_semester_sort_key):
        res_list = by_sem[sem]
        sub_items: List[ReportCardSubject] = []
        sem_credits = 0.0
        sem_weighted = 0.0
        sem_earned = 0.0

        for r in res_list:
            sub = subjects_map.get(r.subject_id)
            code = sub.code if sub else f"SUB{r.subject_id}"
            name = sub.name if sub else "Subject"
            credits = sub.credits if sub else 0.0

            meta = subject_meta.get(r.subject_id, {})
            curr_attempt = meta.get("seq_map", {}).get(r.id, r.attempt or 1)
            is_this_fail = (r.grade_point == 0 or (r.grade or "").strip().upper() in ("F", "AB", "ABSENT", "NC", "E", "Z"))
            is_this_cleared = (not is_this_fail) and (r.grade_point > 0)
            is_arrear_sub = meta.get("had_fail", False) or (curr_attempt > 1) or bool(r.had_arrear)

            sub_items.append(ReportCardSubject(
                code=code,
                name=name,
                credits=credits,
                grade=r.grade,
                grade_point=r.grade_point,
                attempt=curr_attempt,
                total_attempts=meta.get("total_attempts", 1),
                is_arrear=is_arrear_sub,
                is_cleared=is_this_cleared,
                is_failed=is_this_fail,
                failed_in_semester=meta.get("failed_sem"),
                cleared_in_semester=meta.get("cleared_sem"),
                cleared_grade=meta.get("cleared_grd"),
                original_semester=meta.get("original_sem"),
                attempts_history=meta.get("history", [])
            ))

            sem_credits += credits
            sem_weighted += (r.grade_point * credits)
            if r.grade_point > 0:
                sem_earned += credits

        sub_items.sort(key=lambda s: s.code)
        sgpa = round(sem_weighted / sem_credits, 2) if sem_credits > 0 else None

        semesters.append(ReportCardSemester(
            semester=sem,
            subjects=sub_items,
            sgpa=sgpa,
            total_credits=sem_credits,
            earned_credits=sem_earned,
        ))

    # 3. Overall CGPA & Academic Credits Calculation (using best/latest unique subject attempts)
    cgpa_total_credits = 0.0
    cgpa_total_weighted = 0.0
    cgpa_total_earned = 0.0

    for sub_id, att_list in subject_results.items():
        sub = subjects_map.get(sub_id)
        credits = sub.credits if sub else 0.0
        if credits <= 0:
            continue

        # Choose best passing attempt if available, else latest attempt
        passed_attempts = [a for a in att_list if a.grade_point > 0 and (a.grade or "").strip().upper() not in ("F", "AB", "ABSENT", "NC", "E", "Z")]
        if passed_attempts:
            best_att = max(passed_attempts, key=lambda a: a.grade_point)
            cgpa_total_credits += credits
            cgpa_total_weighted += (best_att.grade_point * credits)
            cgpa_total_earned += credits
        else:
            latest_att = att_list[-1]
            cgpa_total_credits += credits
            cgpa_total_weighted += (latest_att.grade_point * credits)

    cgpa = round(cgpa_total_weighted / cgpa_total_credits, 2) if cgpa_total_credits > 0 else None
    cgpa_percentage = round(cgpa * 10, 2) if cgpa is not None else None

    return ReportCardResponse(
        reg_no=student.reg_no,
        name=student.name,
        department=student.department,
        programme=student.programme,
        batch=student.batch,
        section=student.section,
        semesters=semesters,
        arrear_history=arrear_history_list,
        has_arrears=has_any_arrears,
        cgpa=cgpa,
        cgpa_percentage=cgpa_percentage,
        total_credits=cgpa_total_credits,
        earned_credits=cgpa_total_earned,
    )

@app.get("/report-card/lookup-email/{reg_no}")
def lookup_student_email(reg_no: str):
    """
    Returns the student's email for auto-fill when 10-digit reg_no is entered.
    Defaults to {reg_no}@ptuniv.edu.in if student exists or default pattern.
    """
    cleaned_reg = (reg_no or "").strip()
    if not cleaned_reg:
        raise HTTPException(status_code=400, detail="Register Number is required")

    default_email = f"{cleaned_reg}@ptuniv.edu.in"
    found_info = find_student_batch(cleaned_reg)
    if not found_info:
        return {"reg_no": cleaned_reg, "email": default_email, "name": None, "found": False}

    batch_db = get_batch_session(found_info[1])
    try:
        student = batch_db.query(Student).filter(Student.reg_no.ilike(cleaned_reg)).first()
        if student:
            return {
                "reg_no": cleaned_reg,
                "email": student.email or default_email,
                "name": student.name,
                "found": True
            }
        return {"reg_no": cleaned_reg, "email": default_email, "name": None, "found": False}
    finally:
        batch_db.close()

@app.post("/report-card/request-otp", response_model=OTPRequestResponse)
def request_report_card_otp(data: OTPRequestRequest, db: Session = Depends(get_system_db)):
    reg_no = (data.reg_no or "").strip()
    email = (data.email or "").strip().lower()
    if not reg_no or not email:
        raise HTTPException(status_code=400, detail="Reg No and Email are required")

    generic_response = OTPRequestResponse(
        message="If that Reg No and Email match our records, an OTP has been sent to the registered email address.",
        expires_in_seconds=OTP_EXPIRE_MINUTES * 60,
        resend_after_seconds=OTP_RESEND_COOLDOWN_SECONDS,
    )

    found_info = find_student_batch(reg_no)
    if not found_info:
        return generic_response

    batch_db = get_batch_session(found_info[1])
    try:
        student = batch_db.query(Student).filter(Student.reg_no.ilike(reg_no)).first()
        if not student:
            return generic_response
        
        # If student edited their email in student portal, save/update it so OTP and report card verification match
        if email and (not student.email or student.email.strip().lower() != email):
            student.email = email
            batch_db.commit()

        student_name = student.name
    finally:
        batch_db.close()

    now = datetime.datetime.utcnow()
    last_otp = (
        db.query(ReportCardOTP)
        .filter(ReportCardOTP.reg_no.ilike(reg_no))
        .order_by(ReportCardOTP.created_at.desc())
        .first()
    )
    if last_otp and (now - last_otp.created_at).total_seconds() < OTP_RESEND_COOLDOWN_SECONDS:
        wait_left = OTP_RESEND_COOLDOWN_SECONDS - int((now - last_otp.created_at).total_seconds())
        raise HTTPException(status_code=429, detail=f"Please wait {max(wait_left, 1)}s before requesting another OTP")

    hour_ago = now - datetime.timedelta(hours=1)
    recent_count = (
        db.query(ReportCardOTP)
        .filter(ReportCardOTP.reg_no.ilike(reg_no), ReportCardOTP.created_at >= hour_ago)
        .count()
    )
    if recent_count >= OTP_MAX_REQUESTS_PER_HOUR:
        raise HTTPException(status_code=429, detail="Too many OTP requests. Please try again later.")

    db.query(ReportCardOTP).filter(
        ReportCardOTP.reg_no.ilike(reg_no), ReportCardOTP.consumed == False  # noqa: E712
    ).update({"consumed": True}, synchronize_session=False)

    otp = _generate_otp()
    otp_row = ReportCardOTP(
        reg_no=reg_no,
        email=email,
        otp_hash=_hash_otp(otp),
        created_at=now,
        expires_at=now + datetime.timedelta(minutes=OTP_EXPIRE_MINUTES),
        attempts=0,
        consumed=False,
    )
    db.add(otp_row)
    db.commit()

    sent = _send_otp_email(email, student_name, reg_no, otp)
    if not sent:
        db.delete(otp_row)
        db.commit()
        raise HTTPException(
            status_code=503,
            detail="Unable to deliver OTP to your email. Please ensure your email service is configured or try again later."
        )

    return generic_response

@app.post("/report-card/verify-otp", response_model=OTPVerifyResponse)
def verify_report_card_otp(data: OTPVerifyRequest, db: Session = Depends(get_system_db)):
    reg_no = (data.reg_no or "").strip()
    email = (data.email or "").strip().lower()
    otp = (data.otp or "").strip()
    if not reg_no or not email or not otp:
        raise HTTPException(status_code=400, detail="Reg No, Email and OTP are required")
    if not re.fullmatch(r"\d{6}", otp):
        raise HTTPException(status_code=400, detail="Enter the 6-digit OTP sent to your email")

    otp_row = (
        db.query(ReportCardOTP)
        .filter(
            ReportCardOTP.reg_no.ilike(reg_no),
            ReportCardOTP.email == email,
            ReportCardOTP.consumed == False,  # noqa: E712
        )
        .order_by(ReportCardOTP.created_at.desc())
        .first()
    )
    if not otp_row:
        raise HTTPException(status_code=400, detail="No active OTP found for this Reg No/Email. Please request a new one.")

    now = datetime.datetime.utcnow()
    if now > otp_row.expires_at:
        otp_row.consumed = True
        db.commit()
        raise HTTPException(status_code=400, detail="This OTP has expired. Please request a new one.")

    if otp_row.attempts >= OTP_MAX_ATTEMPTS:
        otp_row.consumed = True
        db.commit()
        raise HTTPException(status_code=400, detail="Too many incorrect attempts. Please request a new OTP.")

    if not _verify_otp_hash(otp, otp_row.otp_hash):
        otp_row.attempts += 1
        remaining = max(OTP_MAX_ATTEMPTS - otp_row.attempts, 0)
        if remaining <= 0:
            otp_row.consumed = True
        db.commit()
        detail = "Too many incorrect attempts. Please request a new OTP." if remaining <= 0 \
            else f"Incorrect OTP. {remaining} attempt(s) remaining."
        raise HTTPException(status_code=400, detail=detail)

    otp_row.consumed = True
    db.commit()

    token = _issue_report_card_token(otp_row.reg_no, email)
    return OTPVerifyResponse(access_token=token, expires_in_seconds=REPORT_CARD_TOKEN_EXPIRE_MINUTES * 60)

@app.post("/report-card", response_model=ReportCardResponse)
def get_report_card(
    data: ReportCardRequest,
    subjects_db: Session = Depends(get_subjects_db)
):
    reg_no = (data.reg_no or "").strip()
    email = (data.email or "").strip().lower()
    if not reg_no or not email:
        raise HTTPException(status_code=400, detail="Reg No and Email are required")
    if not data.access_token:
        raise HTTPException(status_code=401, detail="OTP verification required")

    _verify_report_card_token(data.access_token, reg_no, email)

    found_info = find_student_batch(reg_no)
    if not found_info:
        raise HTTPException(status_code=404, detail="No record found for that Reg No and Email combination")

    batch_db = get_batch_session(found_info[1])
    try:
        student = batch_db.query(Student).filter(Student.reg_no.ilike(reg_no)).first()
        if not student or not student.email or student.email.strip().lower() != email:
            raise HTTPException(status_code=404, detail="No record found for that Reg No and Email combination")
        return _build_student_report_card(student, batch_db, subjects_db)
    finally:
        batch_db.close()

@app.get("/admin/report-card/{reg_no}", response_model=ReportCardResponse)
def get_direct_report_card(
    reg_no: str,
    subjects_db: Session = Depends(get_subjects_db),
    user: User = Depends(allow_all)
):
    cleaned_reg = (reg_no or "").strip()
    if not cleaned_reg:
        raise HTTPException(status_code=400, detail="Register Number is required")

    found_info = find_student_batch(cleaned_reg)
    if not found_info:
        raise HTTPException(status_code=404, detail=f"No student found with Register Number '{cleaned_reg}'")

    batch_db = get_batch_session(found_info[1])
    try:
        student = batch_db.query(Student).filter(Student.reg_no.ilike(cleaned_reg)).first()
        if not student:
            raise HTTPException(status_code=404, detail=f"No student found with Register Number '{cleaned_reg}'")
        return _build_student_report_card(student, batch_db, subjects_db)
    finally:
        batch_db.close()

# ==========================================
# STATIC FRONTEND SERVING (Cloud & Public Hosting)
# ==========================================
FRONTEND_DIR = os.path.join(os.path.dirname(BASE_DIR), "frontend")
if os.path.exists(FRONTEND_DIR):
    @app.get("/", include_in_schema=False)
    def serve_frontend_root():
        index_file = os.path.join(FRONTEND_DIR, "Result.html")
        if os.path.exists(index_file):
            return FileResponse(index_file)
        return {"status": "running", "message": "PTU Grade Portal API"}

    @app.get("/Result.html", include_in_schema=False)
    def serve_frontend_html():
        return FileResponse(os.path.join(FRONTEND_DIR, "Result.html"))

    @app.get("/Result.css", include_in_schema=False)
    def serve_frontend_css():
        return FileResponse(os.path.join(FRONTEND_DIR, "Result.css"))

    @app.get("/Result.js", include_in_schema=False)
    def serve_frontend_js():
        return FileResponse(os.path.join(FRONTEND_DIR, "Result.js"))

    @app.get("/Result-motion.js", include_in_schema=False)
    def serve_frontend_motion_js():
        return FileResponse(os.path.join(FRONTEND_DIR, "Result-motion.js"))

# ==========================================
# LIFESPAN & INIT
# ==========================================
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Schema migration check: ensure 'email' column exists on 'users' table
    with system_engine.connect() as conn:
        insp = inspect(system_engine)
        if insp.has_table("users"):
            cols = [c["name"] for c in insp.get_columns("users")]
            if "email" not in cols:
                conn.execute(text("ALTER TABLE users ADD COLUMN email VARCHAR"))
                conn.commit()

    # Initialize System and Subjects DB tables (creates resources table if missing)
    SystemBase.metadata.create_all(bind=system_engine)
    SubjectsBase.metadata.create_all(bind=subjects_engine)

    # Initialize default admin and admin resource in system.db
    sys_db = SystemSessionLocal()
    try:
        admin = sys_db.query(User).filter(User.username == "shakthivel").first()
        if not admin:
            admin = User(
                username="shakthivel",
                email="shakthivel@ptuniv.edu.in",
                hashed_password=hash_password("mK9#vP2$xL8%rQ4!"),
                role="Admin"
            )
            sys_db.add(admin)
            sys_db.commit()
        else:
            if not admin.email:
                admin.email = "shakthivel@ptuniv.edu.in"
            if admin.role.lower() == "admin":
                admin.role = "Admin"
            sys_db.commit()

        # Ensure corresponding Resource exists for admin so pre-existence validation passes
        admin_res = sys_db.query(Resource).filter(
            (func.lower(Resource.name) == "shakthivel") |
            (func.lower(Resource.email) == "shakthivel@ptuniv.edu.in")
        ).first()
        if not admin_res:
            admin_res = Resource(
                name="shakthivel",
                email="shakthivel@ptuniv.edu.in",
                account_type="Admin"
            )
            sys_db.add(admin_res)
            sys_db.commit()
    finally:
        sys_db.close()

    yield

app.router.lifespan_context = lifespan

# ==========================================
# RUN
# ==========================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "Result:app",
        host="0.0.0.0",
        port=8000,
        reload=True,
        reload_includes=["*.py"],
        reload_excludes=["*.db", "*.db-journal", "*.sqlite", "*.sqlite3"],
    )